from flask import (
    Flask,
    request,
    abort,
    send_file,
    render_template,
    redirect,
    flash,
    session,
    url_for,
    jsonify
)

from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import (
    MessageEvent,
    TextMessage,
    TextSendMessage,
    PostbackEvent,
    QuickReply,
    QuickReplyButton,
    MessageAction,
    URIAction,
    TemplateSendMessage,
    ConfirmTemplate
)

from urllib.parse import parse_qs, unquote, quote
import urllib.request as req
import traceback
import os
import csv
import random
import string
from functools import wraps

from io import BytesIO, StringIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.exceptions import RequestEntityTooLarge

from config import LINE_CHANNEL_ACCESS_TOKEN, LINE_CHANNEL_SECRET, headers, BASE_URL

from snkrdunk import (
    search_products,
    getprice,
    get_product_id,
    build_sales_history_url,
    get_prices_by_conditions,
    get_base_conditions,
    get_pro_conditions
)

from flex_messages import (
    create_product_image_grid_messages,
    create_price_flex,
    create_price_flex_carousel,
    create_grade_summary_flex,
    create_history_flex,
    create_market_image_card_flex
)

from exchange import get_jpy_spot_sell
from image_utils import crop_white_border, generate_market_card_image

from models import (
    init_db,
    migrate_db,
    add_card,
    get_all_cards,
    get_dashboard_full_summary,
    get_card_by_id,
    update_card,
    update_card_market_price,
    delete_card,
    mark_card_as_sold,
    mark_card_as_holding,
    create_user,
    get_user_by_username,
    get_user_by_id,
    get_first_user,
    assign_unowned_cards_to_user,
    update_user_password,
    update_user_display_name,
    update_user_line_bind_code,
    get_user_by_line_bind_code,
    get_user_by_line_user_id,
    bind_line_user_to_account,
    unbind_line_user,
    delete_user_account,
    update_user_admin_status,
    update_user_membership_level,

    get_admin_overview_stats,
    get_admin_users,
    count_admin_users,
    get_admin_cards,
    count_admin_cards,
    add_line_log,
    get_admin_line_logs,
    count_admin_line_logs,
    get_line_search_popular_keywords,
    get_line_search_no_result_keywords,
    get_recent_line_search_logs,
    count_recent_line_search_logs,

    add_search_alias,
    get_all_search_aliases,
    count_search_aliases,
    get_search_alias_by_id,
    update_search_alias,
    update_search_alias_active,
    delete_search_alias,
    resolve_search_alias,

    add_search_tag,
    get_all_search_tags,
    get_search_tag_by_id,
    update_search_tag,
    delete_search_tag,
    bulk_import_search_aliases,

    create_price_update_job,
    get_price_update_job,
    get_latest_price_update_job_for_user,
    has_recent_price_update_job,
    has_active_price_update_job,
    mark_price_update_job_running,
    update_price_update_job_progress,
    finish_price_update_job,
    fail_stale_price_update_jobs,
    get_jpy_rate_cache,
    upsert_jpy_rate_cache,
    get_market_price_cache,
    upsert_market_price_cache
)

from calculations import calculate_total_cost

from datetime import datetime, date, timedelta,timezone


app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "cadouka-secret-key")

# 圖片上傳限制：前端會先壓縮，後端再保護性限制大小。
CARD_IMAGE_MAX_MB = 3
app.config["MAX_CONTENT_LENGTH"] = CARD_IMAGE_MAX_MB * 1024 * 1024

LINE_LIFF_ID = os.getenv("LINE_LIFF_ID", "")

init_db()
migrate_db()

line_bot_api = LineBotApi(LINE_CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(LINE_CHANNEL_SECRET)


# 暫存 LINE 使用者搜尋到的商品結果
# key = LINE user_id
# value = 商品列表
user_products = {}


# =========================
# Image Upload Helpers
# =========================

ALLOWED_IMAGE_EXTENSIONS = {"jpg", "jpeg", "png", "webp"}


def is_allowed_card_image(filename):
    filename = filename or ""

    if "." not in filename:
        return False

    extension = filename.rsplit(".", 1)[1].lower().strip()

    return extension in ALLOWED_IMAGE_EXTENSIONS


def upload_card_image_to_cloudinary(uploaded_file, user_id=None):
    """
    使用者上傳圖片優先於預設 image_url。
    圖片實際存在 Cloudinary，資料庫只保存 secure_url。
    """
    if not uploaded_file or not uploaded_file.filename:
        return ""

    if not is_allowed_card_image(uploaded_file.filename):
        raise ValueError("圖片格式不支援，請上傳 JPG、PNG 或 WEBP。")

    cloud_name = os.getenv("CLOUDINARY_CLOUD_NAME", "").strip()
    api_key = os.getenv("CLOUDINARY_API_KEY", "").strip()
    api_secret = os.getenv("CLOUDINARY_API_SECRET", "").strip()

    if not cloud_name or not api_key or not api_secret:
        raise ValueError("圖片上傳功能尚未設定完成，請先設定 Cloudinary 環境變數。")

    try:
        import cloudinary
        import cloudinary.uploader
    except ImportError:
        raise ValueError("伺服器尚未安裝 cloudinary 套件，請在 requirements.txt 加上 cloudinary。")

    cloudinary.config(
        cloud_name=cloud_name,
        api_key=api_key,
        api_secret=api_secret,
        secure=True
    )

    folder_user = f"user_{user_id}" if user_id else "unassigned"

    try:
        result = cloudinary.uploader.upload(
            uploaded_file,
            folder=f"cadouka/cards/{folder_user}",
            resource_type="image",
            use_filename=True,
            unique_filename=True,
            overwrite=False
        )
    except Exception as e:
        print("Cloudinary 圖片上傳失敗：", e)
        traceback.print_exc()
        raise ValueError("圖片上傳失敗，請稍後再試或改用圖片網址。")

    return result.get("secure_url") or result.get("url") or ""


@app.errorhandler(RequestEntityTooLarge)
def handle_uploaded_file_too_large(error):
    flash(f"圖片檔案過大，請上傳 {CARD_IMAGE_MAX_MB}MB 以下圖片。", "warning")
    return redirect(request.referrer or "/cards")


def get_optimized_image_url(image_url, preset="detail"):
    """
    圖片顯示規則：
    1. 使用者自己上傳到 Cloudinary 的圖片：使用 Cloudinary transformation 加速。
    2. SNKRDUNK / 外部圖片：維持原本 /crop-image 後端裁白邊設定。
    """
    if not image_url:
        return ""

    image_url = str(image_url).strip()

    if not image_url:
        return ""

    is_cloudinary = "res.cloudinary.com" in image_url and "/image/upload/" in image_url

    # 非 Cloudinary 圖片，多半是 SNKRDUNK 圖片，維持原本裁白邊邏輯。
    if not is_cloudinary:
        return f"/crop-image?url={quote(image_url, safe='')}"

    presets = {
        "thumb": "f_auto,q_auto,w_180,h_240,c_fit",
        "list": "f_auto,q_auto,w_220,h_300,c_fit",
        "detail": "f_auto,q_auto,w_800,c_fit",
        "large": "f_auto,q_auto,w_1200,c_fit"
    }

    transformation = presets.get(preset, presets["detail"])

    # 避免已經處理過的網址被重複插入 transformation。
    upload_marker = "/image/upload/"
    before, after = image_url.split(upload_marker, 1)
    first_segment = after.split("/", 1)[0]

    if first_segment.startswith("f_auto") or first_segment.startswith("q_auto") or "w_" in first_segment:
        return image_url

    return f"{before}{upload_marker}{transformation}/{after}"


# =========================
# Condition Order Settings
# =========================

# Free users only see PSA conditions.
BASE_CONDITION_ORDER = ["PSA10", "PSA9", "PSA8以下"]

# Pro/Admin users see extra raw-card conditions after PSA conditions.
PRO_CONDITION_ORDER = ["PSA10", "PSA9", "PSA8以下", "A", "B"]


def get_base_conditions():
    return BASE_CONDITION_ORDER


def get_pro_conditions():
    return PRO_CONDITION_ORDER


# =========================
# Auth Helpers
# =========================

def current_user_id():
    return session.get("user_id")


def current_user():
    user_id = current_user_id()

    if not user_id:
        return None

    return get_user_by_id(user_id)


def generate_line_bind_code():
    characters = string.ascii_uppercase + string.digits

    for _ in range(20):
        bind_code = "".join(random.choice(characters) for _ in range(6))

        existing_user = get_user_by_line_bind_code(bind_code)

        if not existing_user:
            return bind_code

    return "".join(random.choice(characters) for _ in range(10))


def login_required(view_func):
    @wraps(view_func)
    def wrapped_view(*args, **kwargs):
        if not current_user_id():
            flash("請先登入", "warning")
            return redirect(url_for("login_page", next=request.full_path))

        return view_func(*args, **kwargs)

    return wrapped_view

def admin_required(view_func):
    @wraps(view_func)
    def wrapped_view(*args, **kwargs):
        user = current_user()

        if not user:
            flash("請先登入", "warning")
            return redirect("/login")

        try:
            is_admin = user["is_admin"] or 0
        except:
            is_admin = 0

        if int(is_admin) != 1:
            flash("你沒有權限進入後台", "warning")
            return redirect("/dashboard")

        return view_func(*args, **kwargs)

    return wrapped_view

def get_membership_level(user):
    if not user:
        return "free"

    try:
        membership_level = user["membership_level"] or "free"
    except:
        membership_level = "free"

    return membership_level


def is_pro_user(user):
    if not user:
        return False

    try:
        is_admin = int(user["is_admin"] or 0)
    except:
        is_admin = 0

    membership_level = get_membership_level(user)

    return is_admin == 1 or membership_level == "pro"

@app.context_processor
def inject_current_user():
    user = current_user()

    return {
        "current_user": user,
        "line_liff_id": LINE_LIFF_ID,
        "membership_level": get_membership_level(user),
        "is_pro": is_pro_user(user),
        "get_optimized_image_url": get_optimized_image_url
    }

# =========================
# LINE Quick Reply Helpers
# =========================

def get_base_url():
    base_url = BASE_URL.strip().rstrip("/")

    if not base_url.startswith("http://") and not base_url.startswith("https://"):
        base_url = "https://" + base_url

    return base_url


def create_main_quick_reply():
    base_url = get_base_url()

    return QuickReply(
        items=[
            QuickReplyButton(
                action=URIAction(
                    label="卡牌倉庫",
                    uri=f"{base_url}/cards"
                )
            ),
            QuickReplyButton(
                action=MessageAction(
                    label="綁定狀態",
                    text="綁定狀態"
                )
            ),
            QuickReplyButton(
                action=MessageAction(
                    label="使用教學",
                    text="使用教學"
                )
            ),
            QuickReplyButton(
                action=MessageAction(
                    label="解除綁定",
                    text="解除綁定"
                )
            )
        ]
    )

def create_tutorial_quick_reply():
    return QuickReply(
        items=[
            QuickReplyButton(
                action=MessageAction(
                    label="如何綁定",
                    text="教學：如何綁定"
                )
            ),
            QuickReplyButton(
                action=MessageAction(
                    label="如何查價",
                    text="教學：如何查價"
                )
            ),
            QuickReplyButton(
                action=MessageAction(
                    label="加入倉庫",
                    text="教學：加入倉庫"
                )
            ),
            QuickReplyButton(
                action=MessageAction(
                    label="常用指令",
                    text="教學：常用指令"
                )
            )
        ]
    )

def create_unbound_quick_reply():
    base_url = get_base_url()

    return QuickReply(
        items=[
            QuickReplyButton(
                action=URIAction(
                    label="開啟 Cadouka",
                    uri=f"{base_url}/profile"
                )
            ),
            QuickReplyButton(
                action=MessageAction(
                    label="如何綁定",
                    text="教學：如何綁定"
                )
            ),
            QuickReplyButton(
                action=MessageAction(
                    label="綁定狀態",
                    text="綁定狀態"
                )
            )
        ]
    )

# =========================
# LINE Log Helper
# =========================

def safe_add_line_log(
    line_user_id=None,
    action="",
    result="",
    message="",
    raw_keyword="",
    resolved_keyword="",
    product_count=None
):
    try:
        user = get_user_by_line_user_id(line_user_id) if line_user_id else None
        user_id = user["id"] if user else None

        add_line_log(
            line_user_id=line_user_id,
            user_id=user_id,
            action=action,
            result=result,
            message=message,
            raw_keyword=raw_keyword,
            resolved_keyword=resolved_keyword,
            product_count=product_count
        )
    except Exception as e:
        print("LINE log 寫入失敗：", e)
        traceback.print_exc()

# =========================
# Calculation Helpers
# =========================

def calculate_holding_days_for_card(card):
    """
    持有中：今天 - 購入日期
    已售出：售出日期 - 購入日期
    沒有購入日期：回傳 "-"
    """
    buy_date_text = card.get("buy_date") if isinstance(card, dict) else card["buy_date"]

    if not buy_date_text:
        return "-"

    try:
        buy_date = datetime.strptime(buy_date_text, "%Y-%m-%d").date()
    except:
        return "-"

    status = card.get("status") if isinstance(card, dict) else card["status"]

    if status == "sold":
        sell_date_text = card.get("sell_date") if isinstance(card, dict) else card["sell_date"]

        if sell_date_text:
            try:
                end_date = datetime.strptime(sell_date_text, "%Y-%m-%d").date()
            except:
                end_date = date.today()
        else:
            end_date = date.today()
    else:
        end_date = date.today()

    days = (end_date - buy_date).days

    if days < 0:
        days = 0

    return f"{days} 天"


def calculate_unrealized_by_buy_price(current_market_price, buy_price):
    """
    未實現損益 = 目前市價 - 購入價格
    未實現 ROI = 未實現損益 / 購入價格 * 100%
    """
    unrealized_profit = current_market_price - buy_price

    if buy_price == 0:
        roi = 0
    else:
        roi = (unrealized_profit / buy_price) * 100

    return unrealized_profit, roi


def calculate_realized_by_buy_price(net_revenue, buy_price):
    """
    已實現損益 = 實際收入 - 購入價格
    已實現 ROI = 已實現損益 / 購入價格 * 100%
    """
    realized_profit = net_revenue - buy_price

    if buy_price == 0:
        realized_roi = 0
    else:
        realized_roi = (realized_profit / buy_price) * 100

    return realized_profit, realized_roi


def format_jpy_price_for_card(price):
    """
    LINE 後端新增卡牌時，用來把 SNKRDUNK 成交價轉成數字。
    """
    try:
        return int(float(str(price).replace(",", "")))
    except:
        return None


def calculate_market_price_from_prices(prices, jpy_rate):
    """
    用 PSA10 成交紀錄平均價 * 日圓匯率，算出目前市價。
    沒有資料或匯率失敗時回傳 0。
    """
    if not prices or not jpy_rate:
        return 0

    valid_prices = []

    for p in prices:
        jpy_price = format_jpy_price_for_card(p.get("price"))

        if jpy_price is not None:
            valid_prices.append(jpy_price)

    if not valid_prices:
        return 0

    average_jpy = round(sum(valid_prices) / len(valid_prices))

    return round(average_jpy * jpy_rate)

def parse_jpy_price(price):
    try:
        return int(float(str(price).replace(",", "").replace("¥", "").strip()))
    except:
        return None


def is_recent_24h_sale(date_text):
    """
    SNKRDUNK 回傳常見格式：
    8分前、2時間前、12時間前、1日前、26/06/05
    第一版只把 分前 / 時間前 / 剛剛 算進 24h。
    """
    if not date_text:
        return False

    date_text = str(date_text)

    if "分前" in date_text:
        return True

    if "時間前" in date_text:
        return True

    if "剛剛" in date_text:
        return True

    return False


def format_number_for_analysis(value):
    try:
        return f"{float(value):,.0f}"
    except:
        return "0"


def format_percent_for_analysis(value, digits=1, show_sign=False):
    try:
        number = float(value)
    except:
        number = 0

    sign = ""

    if show_sign and number > 0:
        sign = "+"

    return f"{sign}{number:.{digits}f}%"


def get_card_value(card, key, default=None):
    if not card:
        return default

    try:
        if isinstance(card, dict):
            return card.get(key, default)

        value = card[key]
        return value if value is not None else default
    except:
        return default


def get_analysis_grade(card, grade_order):
    grade = str(get_card_value(card, "grade", "") or "").strip()

    if grade in grade_order:
        return grade

    return "PSA10"


def parse_holding_days_number(holding_days_text):
    if holding_days_text is None:
        return None

    text = str(holding_days_text).strip()

    if not text or text == "-":
        return None

    try:
        return int(text.replace("天", "").strip())
    except:
        return None


def make_advanced_indicator(key, title, value, status, description, level="neutral"):
    return {
        "key": key,
        "title": title,
        "value": value,
        "status": status,
        "description": description,
        "level": level
    }


def summarize_price_list(prices, jpy_rate):
    valid_prices = []
    latest_jpy = None

    for item in prices or []:
        jpy_price = parse_jpy_price(item.get("price"))

        if jpy_price is not None:
            valid_prices.append(jpy_price)

            if latest_jpy is None:
                latest_jpy = jpy_price

    if not valid_prices:
        return {
            "has_data": False,
            "count_24h": 0,
            "sale_count": 0,
            "latest_jpy": 0,
            "highest_jpy": 0,
            "average_jpy": 0,
            "lowest_jpy": 0,
            "latest_twd": 0,
            "highest_twd": 0,
            "average_twd": 0,
            "lowest_twd": 0
        }

    highest_jpy = max(valid_prices)
    lowest_jpy = min(valid_prices)
    average_jpy = round(sum(valid_prices) / len(valid_prices))

    count_24h = 0

    for item in prices or []:
        if is_recent_24h_sale(item.get("date")):
            count_24h += 1

    return {
        "has_data": True,
        "count_24h": count_24h,
        "sale_count": len(valid_prices),

        "latest_jpy": latest_jpy or 0,
        "highest_jpy": highest_jpy,
        "average_jpy": average_jpy,
        "lowest_jpy": lowest_jpy,

        "latest_twd": round(latest_jpy * jpy_rate) if latest_jpy and jpy_rate else 0,
        "highest_twd": round(highest_jpy * jpy_rate) if jpy_rate else 0,
        "average_twd": round(average_jpy * jpy_rate) if jpy_rate else 0,
        "lowest_twd": round(lowest_jpy * jpy_rate) if jpy_rate else 0
    }


def build_market_advanced_indicators(card, analysis_grade, grade_summary):
    indicators = []

    if not grade_summary or not grade_summary.get("has_data"):
        return [
            make_advanced_indicator(
                "price_position",
                "價格位置",
                "資料不足",
                "無法判斷",
                f"{analysis_grade} 目前沒有足夠成交資料。",
                "muted"
            ),
            make_advanced_indicator(
                "price_volatility",
                "成交價波動",
                "資料不足",
                "無法判斷",
                f"{analysis_grade} 目前沒有足夠成交資料。",
                "muted"
            ),
            make_advanced_indicator(
                "market_heat",
                "成交熱度",
                "低",
                "24h 0 筆",
                "近期成交量偏少，流動性需要多觀察。",
                "muted"
            ),
            make_advanced_indicator(
                "safety_margin",
                "安全邊際",
                "資料不足",
                "無法判斷",
                "請先填入購入價格，才能計算安全邊際。",
                "muted"
            ),
            make_advanced_indicator(
                "capital_efficiency",
                "資金效率",
                "資料不足",
                "無法判斷",
                "請先填入購入日期與購入價格，才能計算資金效率。",
                "muted"
            )
        ]

    latest_jpy = grade_summary.get("latest_jpy") or 0
    highest_jpy = grade_summary.get("highest_jpy") or 0
    average_jpy = grade_summary.get("average_jpy") or 0
    lowest_jpy = grade_summary.get("lowest_jpy") or 0
    average_twd = grade_summary.get("average_twd") or 0
    count_24h = grade_summary.get("count_24h") or 0

    # 1. 價格位置
    if latest_jpy > 0 and highest_jpy > lowest_jpy:
        price_position = ((latest_jpy - lowest_jpy) / (highest_jpy - lowest_jpy)) * 100
        price_position = max(0, min(100, price_position))

        if price_position <= 30:
            position_status = "偏低區"
            position_level = "good"
        elif price_position <= 70:
            position_status = "合理區"
            position_level = "neutral"
        else:
            position_status = "偏高區"
            position_level = "warning"

        indicators.append(
            make_advanced_indicator(
                "price_position",
                "價格位置",
                format_percent_for_analysis(price_position),
                position_status,
                f"最新成交 ¥{format_number_for_analysis(latest_jpy)}，位於近期高低區間的 {format_percent_for_analysis(price_position)}。",
                position_level
            )
        )
    else:
        indicators.append(
            make_advanced_indicator(
                "price_position",
                "價格位置",
                "區間不足",
                "無法判斷",
                "近期最高價與最低價太接近，暫時無法判斷價格位置。",
                "muted"
            )
        )

    # 2. 成交價波動
    if average_jpy > 0 and highest_jpy >= lowest_jpy:
        volatility = ((highest_jpy - lowest_jpy) / average_jpy) * 100

        if volatility <= 10:
            volatility_status = "穩定"
            volatility_level = "good"
        elif volatility <= 25:
            volatility_status = "中等波動"
            volatility_level = "neutral"
        else:
            volatility_status = "波動偏大"
            volatility_level = "warning"

        indicators.append(
            make_advanced_indicator(
                "price_volatility",
                "成交價波動",
                format_percent_for_analysis(volatility),
                volatility_status,
                f"近期最高 ¥{format_number_for_analysis(highest_jpy)}，最低 ¥{format_number_for_analysis(lowest_jpy)}。",
                volatility_level
            )
        )
    else:
        indicators.append(
            make_advanced_indicator(
                "price_volatility",
                "成交價波動",
                "資料不足",
                "無法判斷",
                "目前成交資料不足，暫時無法計算成交價波動。",
                "muted"
            )
        )

    # 3. 成交熱度
    if count_24h >= 10:
        heat_value = "高"
        heat_level = "good"
        heat_desc = "24h 成交量活躍，流動性相對較好。"
    elif count_24h >= 3:
        heat_value = "中"
        heat_level = "neutral"
        heat_desc = "24h 有一定成交量，流動性普通。"
    else:
        heat_value = "低"
        heat_level = "muted"
        heat_desc = "24h 成交量偏少，買賣可能需要更多時間。"

    indicators.append(
        make_advanced_indicator(
            "market_heat",
            "成交熱度",
            heat_value,
            f"24h {count_24h} 筆",
            heat_desc,
            heat_level
        )
    )

    # 4. 安全邊際
    total_cost = get_card_value(card, "total_cost", 0) or 0
    buy_price = get_card_value(card, "buy_price", 0) or 0

    try:
        cost_basis = float(total_cost or buy_price or 0)
    except:
        cost_basis = 0

    if cost_basis > 0 and average_twd > 0:
        safety_margin = ((average_twd - cost_basis) / average_twd) * 100

        if safety_margin >= 15:
            margin_status = "安全邊際佳"
            margin_level = "good"
        elif safety_margin >= 0:
            margin_status = "接近均價"
            margin_level = "neutral"
        else:
            margin_status = "成本偏高"
            margin_level = "warning"

        if safety_margin >= 0:
            margin_desc = (
                f"你的成本 NT${format_number_for_analysis(cost_basis)}，"
                f"低於近期平均 NT${format_number_for_analysis(average_twd)}。"
            )
        else:
            margin_desc = (
                f"你的成本 NT${format_number_for_analysis(cost_basis)}，"
                f"高於近期平均 NT${format_number_for_analysis(average_twd)}。"
            )

        indicators.append(
            make_advanced_indicator(
                "safety_margin",
                "安全邊際",
                format_percent_for_analysis(safety_margin, show_sign=True),
                margin_status,
                margin_desc,
                margin_level
            )
        )
    else:
        indicators.append(
            make_advanced_indicator(
                "safety_margin",
                "安全邊際",
                "資料不足",
                "無法判斷",
                "請先填入購入價格，才能和近期市場平均比較。",
                "muted"
            )
        )

    # 5. 資金效率，不做年化，只看每日 ROI
    status = str(get_card_value(card, "status", "holding") or "holding")

    if status == "sold":
        roi = get_card_value(card, "realized_roi", None)
    else:
        roi = get_card_value(card, "roi", None)

    holding_days = parse_holding_days_number(get_card_value(card, "holding_days", None))
    buy_date_text = get_card_value(card, "buy_date", "")

    try:
        roi_value = float(roi)
    except:
        roi_value = None

    if roi_value is not None and buy_date_text and holding_days is not None and cost_basis > 0:
        days_for_calc = holding_days if holding_days > 0 else 1
        daily_roi = roi_value / days_for_calc

        if daily_roi >= 0.2:
            efficiency_status = "高"
            efficiency_level = "good"
        elif daily_roi >= 0.05:
            efficiency_status = "中"
            efficiency_level = "neutral"
        elif daily_roi >= 0:
            efficiency_status = "低"
            efficiency_level = "muted"
        else:
            efficiency_status = "虧損中"
            efficiency_level = "warning"

        if holding_days == 0:
            efficiency_desc = f"持有未滿 1 天，ROI {format_percent_for_analysis(roi_value, digits=2, show_sign=True)}。"
        else:
            efficiency_desc = f"持有 {holding_days} 天，ROI {format_percent_for_analysis(roi_value, digits=2, show_sign=True)}。"

        indicators.append(
            make_advanced_indicator(
                "capital_efficiency",
                "資金效率",
                f"每日 {format_percent_for_analysis(daily_roi, digits=2, show_sign=True)}",
                efficiency_status,
                efficiency_desc,
                efficiency_level
            )
        )
    else:
        indicators.append(
            make_advanced_indicator(
                "capital_efficiency",
                "資金效率",
                "資料不足",
                "無法判斷",
                "請先填入購入日期與購入價格，才能計算每日 ROI。",
                "muted"
            )
        )

    return indicators


def build_pro_market_summary(product_url, card=None):
    if not product_url:
        return {
            "available": False,
            "message": "這張卡尚未設定商品網址，無法取得市場分析資料。",
            "jpy_rate": None,
            "analysis_grade": None,
            "advanced_indicators": [],
            "grades": []
        }

    product_id = get_product_id(product_url)

    if not product_id:
        return {
            "available": False,
            "message": "商品網址格式異常，無法取得市場分析資料。",
            "jpy_rate": None,
            "analysis_grade": None,
            "advanced_indicators": [],
            "grades": []
        }

    try:
        grade_order = get_pro_conditions()

        prices_by_conditions = get_prices_by_conditions(
            product_id,
            conditions=grade_order
        )

        jpy_rate = get_jpy_spot_sell()

        grades = []
        summaries_by_grade = {}

        for grade in grade_order:
            prices = prices_by_conditions.get(grade, [])
            summary = summarize_price_list(prices, jpy_rate)
            summaries_by_grade[grade] = summary

            grades.append({
                "grade": grade,
                **summary
            })

        analysis_grade = get_analysis_grade(card, grade_order)
        analysis_summary = summaries_by_grade.get(analysis_grade)

        advanced_indicators = build_market_advanced_indicators(
            card,
            analysis_grade,
            analysis_summary
        )

        return {
            "available": True,
            "message": "",
            "jpy_rate": jpy_rate,
            "analysis_grade": analysis_grade,
            "advanced_indicators": advanced_indicators,
            "grades": grades
        }

    except Exception as e:
        print("市場分析錯誤：", e)
        traceback.print_exc()

        return {
            "available": False,
            "message": "取得市場分析資料時發生錯誤，請稍後再試。",
            "jpy_rate": None,
            "analysis_grade": None,
            "advanced_indicators": [],
            "grades": []
        }

def build_pro_history_data(card, selected_grade):
    product_url = card.get("product_url") or ""

    if not product_url:
        return {
            "available": False,
            "message": "這張卡尚未設定商品網址，無法取得歷史成交資料。",
            "grade": selected_grade,
            "jpy_rate": None,
            "prices": []
        }

    product_id = get_product_id(product_url)

    if not product_id:
        return {
            "available": False,
            "message": "商品網址格式異常，無法取得歷史成交資料。",
            "grade": selected_grade,
            "jpy_rate": None,
            "prices": []
        }

    try:
        price_url = build_sales_history_url(
            product_id,
            condition=selected_grade,
            page=1,
            per_page=20
        )

        prices = getprice(price_url)
        jpy_rate = get_jpy_spot_sell()

        history_prices = []

        for item in (prices or [])[:10]:
            jpy_price = parse_jpy_price(item.get("price"))

            history_prices.append({
                "date": item.get("date") or "-",
                "grade": item.get("condition") or selected_grade,
                "price_jpy": jpy_price or 0,
                "price_twd": round(jpy_price * jpy_rate) if jpy_price and jpy_rate else 0
            })

        return {
            "available": True,
            "message": "",
            "grade": selected_grade,
            "jpy_rate": jpy_rate,
            "prices": history_prices
        }

    except Exception as e:
        print("Pro 歷史成交錯誤：", e)
        traceback.print_exc()

        return {
            "available": False,
            "message": "取得歷史成交資料時發生錯誤，請稍後再試。",
            "grade": selected_grade,
            "jpy_rate": None,
            "prices": []
        }


def get_market_price_by_product_url(product_url):
    """
    用 SNKRDUNK 商品網址抓最新 PSA10 平均成交價，並換算成台幣。
    """
    if not product_url:
        return 0

    product_id = get_product_id(product_url)

    if not product_id:
        return 0

    price_url = build_sales_history_url(product_id)

    prices = getprice(price_url)
    jpy_rate = get_jpy_spot_sell()

    current_market_price = calculate_market_price_from_prices(
        prices,
        jpy_rate
    )

    return current_market_price

def get_taiwan_now_text():
    taiwan_time = datetime.now(timezone(timedelta(hours=8)))
    return taiwan_time.strftime("%Y-%m-%d %H:%M:%S")


def should_skip_price_update(price_updated_at, cooldown_hours=6):
    """
    如果這張卡在 cooldown_hours 小時內更新過，就略過。
    """
    if not price_updated_at:
        return False

    try:
        updated_text = str(price_updated_at).split(".")[0]
        updated_at = datetime.strptime(updated_text, "%Y-%m-%d %H:%M:%S")
        now = datetime.now(timezone(timedelta(hours=8))).replace(tzinfo=None)

        diff = now - updated_at

        return diff < timedelta(hours=cooldown_hours)
    except:
        return False


# =========================
# Web-only Batch Price Update Helpers
# =========================

PRICE_UPDATE_COOLDOWN_HOURS = 6
MARKET_PRICE_CACHE_HOURS = 6
JPY_RATE_CACHE_HOURS = 6
PRICE_UPDATE_BATCH_SIZE = int(os.getenv("PRICE_UPDATE_BATCH_SIZE", "2"))
PRICE_UPDATE_STALE_MINUTES = 10
PRICE_UPDATE_LOCK_MINUTES = 10


def parse_update_datetime(value):
    if not value:
        return None

    if isinstance(value, datetime):
        return value.replace(tzinfo=None)

    text = str(value).strip()

    if not text:
        return None

    text = text.replace("T", " ").split(".")[0]

    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"]:
        try:
            return datetime.strptime(text[:19], fmt)
        except:
            continue

    return None


def is_cache_valid(updated_at, cache_hours=6):
    updated_dt = parse_update_datetime(updated_at)

    if not updated_dt:
        return False

    now = datetime.now(timezone(timedelta(hours=8))).replace(tzinfo=None)
    return (now - updated_dt) < timedelta(hours=cache_hours)


def get_cached_jpy_spot_sell():
    """
    日圓匯率 6 小時快取。
    抓不到新匯率時，如果資料庫裡有舊匯率，先用舊值避免更新整批失敗。
    """
    cache = get_jpy_rate_cache()

    if cache:
        try:
            cached_rate = float(get_card_value(cache, "rate", 0) or 0)
        except:
            cached_rate = 0

        if cached_rate > 0 and is_cache_valid(
            get_card_value(cache, "updated_at", ""),
            cache_hours=JPY_RATE_CACHE_HOURS
        ):
            return cached_rate
    else:
        cached_rate = 0

    try:
        fresh_rate = float(get_jpy_spot_sell() or 0)
    except Exception as e:
        print("日圓匯率取得失敗：", e)
        fresh_rate = 0

    if fresh_rate > 0:
        upsert_jpy_rate_cache(
            rate=fresh_rate,
            updated_at=get_taiwan_now_text()
        )
        return fresh_rate

    if cached_rate > 0:
        return cached_rate

    return 0


def normalize_market_grade_for_card(card):
    grade = str(get_card_value(card, "grade", "") or "").strip()
    allowed_grades = get_pro_conditions()

    if grade in allowed_grades:
        return grade

    return "PSA10"


def calculate_average_jpy_from_prices(prices):
    valid_prices = []

    for item in prices or []:
        jpy_price = parse_jpy_price(item.get("price"))

        if jpy_price is not None:
            valid_prices.append(jpy_price)

    if not valid_prices:
        return 0

    return round(sum(valid_prices) / len(valid_prices))


def get_market_price_by_product_url_cached(product_url, grade="PSA10"):
    """
    商品市價快取：同 product_id + grade 在 6 小時內只抓一次 SNKRDUNK。
    回傳 (market_price_twd, source_text)。
    """
    if not product_url:
        return 0, "no_product_url"

    product_id = get_product_id(product_url)

    if not product_id:
        return 0, "invalid_product_url"

    grade = grade or "PSA10"
    cache = get_market_price_cache(product_id, grade)

    if cache and is_cache_valid(
        get_card_value(cache, "updated_at", ""),
        cache_hours=MARKET_PRICE_CACHE_HOURS
    ):
        try:
            cached_price = round(float(get_card_value(cache, "market_price_twd", 0) or 0))
        except:
            cached_price = 0

        return cached_price, "cache"

    jpy_rate = get_cached_jpy_spot_sell()

    if not jpy_rate:
        return 0, "no_jpy_rate"

    price_url = build_sales_history_url(
        product_id,
        condition=grade,
        page=1,
        per_page=20
    )

    prices = getprice(price_url)
    market_price_twd = calculate_market_price_from_prices(prices, jpy_rate)
    average_jpy = calculate_average_jpy_from_prices(prices)

    upsert_market_price_cache(
        product_id=product_id,
        grade=grade,
        market_price_twd=market_price_twd,
        average_jpy=average_jpy,
        jpy_rate=jpy_rate,
        updated_at=get_taiwan_now_text(),
        source="SNKRDUNK"
    )

    return market_price_twd, "external"


def collect_price_update_card_ids(user_id, mode="all", card_id=None):
    """
    找出本次需要更新的卡牌。
    全部更新只挑：持有中、有商品網址、6 小時內未更新過。
    """
    if mode == "single":
        card = get_card_by_id(card_id, user_id=user_id)

        if not card:
            return []

        card_dict = dict(card)

        if card_dict.get("status") != "holding":
            return []

        return [int(card_id)]

    cards = get_all_cards(
        status="holding",
        keyword=None,
        sort=None,
        user_id=user_id
    )

    card_ids = []

    for card in cards or []:
        card_dict = dict(card)

        if not card_dict.get("product_url"):
            continue

        if should_skip_price_update(
            card_dict.get("price_updated_at"),
            cooldown_hours=PRICE_UPDATE_COOLDOWN_HOURS
        ):
            continue

        card_ids.append(int(card_dict.get("id")))

    return card_ids


def process_one_card_market_price_update(card_id, user_id):
    """
    實際更新單張卡。
    這個 function 會被 AJAX 分批呼叫使用，不需要背景 worker。
    """
    try:
        card = get_card_by_id(card_id, user_id=user_id)

        if not card:
            return {
                "status": "failed",
                "card_id": card_id,
                "message": "找不到卡牌"
            }

        card_dict = dict(card)

        if card_dict.get("status") != "holding":
            return {
                "status": "skipped",
                "card_id": card_id,
                "message": "已售出卡牌略過"
            }

        product_url = card_dict.get("product_url") or ""

        if not product_url:
            return {
                "status": "skipped",
                "card_id": card_id,
                "message": "沒有商品網址"
            }

        if should_skip_price_update(
            card_dict.get("price_updated_at"),
            cooldown_hours=PRICE_UPDATE_COOLDOWN_HOURS
        ):
            return {
                "status": "skipped",
                "card_id": card_id,
                "message": "6 小時內已更新過"
            }

        grade = normalize_market_grade_for_card(card_dict)
        current_market_price, source = get_market_price_by_product_url_cached(
            product_url,
            grade=grade
        )

        if not current_market_price:
            return {
                "status": "failed",
                "card_id": card_id,
                "message": "查無成交價或匯率失敗"
            }

        buy_price = float(card_dict.get("buy_price") or 0)
        unrealized_profit, roi = calculate_unrealized_by_buy_price(
            current_market_price,
            buy_price
        )

        update_card_market_price(
            card_id,
            current_market_price,
            unrealized_profit,
            roi,
            get_taiwan_now_text(),
            user_id=user_id
        )

        return {
            "status": "updated",
            "card_id": card_id,
            "message": "已更新" if source == "external" else "已用快取更新",
            "market_price": current_market_price,
            "source": source
        }

    except Exception as e:
        print("單張分批市價更新失敗：", e)
        traceback.print_exc()

        return {
            "status": "failed",
            "card_id": card_id,
            "message": "更新失敗"
        }

# =========================
# Auth Routes
# =========================

@app.route("/")
def home():
    if current_user_id():
        return redirect("/dashboard")

    return redirect("/login")


@app.route("/register", methods=["GET", "POST"])
def register_page():
    if current_user_id():
        return redirect("/dashboard")

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        confirm_password = request.form.get("confirm_password", "").strip()
        agree_terms = request.form.get("agree_terms")

        if not username or not password:
            flash("請輸入帳號與密碼", "warning")
            return redirect("/register")

        if len(password) < 6:
            flash("密碼至少需要 6 碼", "warning")
            return redirect("/register")

        if password != confirm_password:
            flash("兩次輸入的密碼不一致", "warning")
            return redirect("/register")

        if agree_terms != "1":
            flash("請先閱讀並同意服務條款與隱私權政策", "warning")
            return redirect("/register")

        existing_user = get_user_by_username(username)

        if existing_user:
            flash("這個帳號已經被使用", "warning")
            return redirect("/register")

        first_user_before_create = get_first_user()

        password_hash = generate_password_hash(password)
        accepted_at = get_taiwan_now_text()

        create_user(
            username,
            password_hash,
            terms_accepted_at=accepted_at,
            privacy_accepted_at=accepted_at
        )

        user = get_user_by_username(username)

        session["user_id"] = user["id"]
        session["username"] = user["username"]

        # 如果這是第一個註冊者，把帳號系統上線前的舊資料歸給他
        if first_user_before_create is None:
            assign_unowned_cards_to_user(user["id"])

        flash("註冊成功，已自動登入", "success")
        return redirect("/dashboard")

    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login_page():
    if current_user_id():
        return redirect("/dashboard")

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        next_url = request.form.get("next", "").strip()

        user = get_user_by_username(username)

        if not user:
            flash("帳號或密碼錯誤", "warning")
            return redirect("/login")

        if not check_password_hash(user["password_hash"], password):
            flash("帳號或密碼錯誤", "warning")
            return redirect("/login")

        session["user_id"] = user["id"]
        session["username"] = user["username"]

        flash("登入成功", "success")

        if next_url and next_url.startswith("/"):
            return redirect(next_url)

        return redirect("/dashboard")

    next_url = request.args.get("next", "")
    return render_template("login.html", next_url=next_url)


@app.route("/logout")
def logout_page():
    session.clear()
    flash("已登出", "success")
    return redirect("/login")


@app.route("/profile")
@login_required
def profile_page():
    user = current_user()

    if not user:
        flash("請先登入", "warning")
        return redirect("/login")

    return render_template("profile.html", user=user)


@app.route("/profile/update-display-name", methods=["POST"])
@login_required
def update_display_name_page():
    user = current_user()

    if not user:
        flash("請先登入", "warning")
        return redirect("/login")

    display_name = request.form.get("display_name", "").strip()

    if len(display_name) > 30:
        flash("暱稱最多 30 個字", "warning")
        return redirect("/profile")

    update_user_display_name(user["id"], display_name)

    if display_name:
        flash("暱稱已更新", "success")
    else:
        flash("已清除暱稱", "success")

    return redirect("/profile")


@app.route("/profile/generate-line-bind-code", methods=["POST"])
@login_required
def generate_line_bind_code_page():
    user = current_user()

    if not user:
        flash("請先登入", "warning")
        return redirect("/login")

    bind_code = generate_line_bind_code()


    taiwan_now = datetime.now(timezone(timedelta(hours=8)))
    expires_at = (taiwan_now + timedelta(minutes=30)).strftime("%Y-%m-%d %H:%M:%S")

    update_user_line_bind_code(user["id"], bind_code, expires_at)

    flash("LINE 綁定連結已產生，請在 30 分鐘內完成綁定", "success")
    return redirect("/profile")


@app.route("/profile/change-password", methods=["POST"])
@login_required
def change_password_page():
    user = current_user()

    if not user:
        flash("請先登入", "warning")
        return redirect("/login")

    current_password = request.form.get("current_password", "").strip()
    new_password = request.form.get("new_password", "").strip()
    confirm_password = request.form.get("confirm_password", "").strip()

    if not current_password or not new_password or not confirm_password:
        flash("請完整填寫密碼欄位", "warning")
        return redirect("/profile")

    if not check_password_hash(user["password_hash"], current_password):
        flash("目前密碼錯誤", "warning")
        return redirect("/profile")

    if len(new_password) < 6:
        flash("新密碼至少需要 6 碼", "warning")
        return redirect("/profile")

    if new_password != confirm_password:
        flash("兩次輸入的新密碼不一致", "warning")
        return redirect("/profile")

    password_hash = generate_password_hash(new_password)
    update_user_password(user["id"], password_hash)

    flash("密碼已成功更新", "success")
    return redirect("/profile")

@app.route("/profile/delete-account", methods=["POST"])
@login_required
def delete_account_page():
    user = current_user()

    if not user:
        flash("請先登入", "warning")
        return redirect("/login")

    user_id = user["id"]

    delete_user_account(user_id)

    session.clear()

    flash("帳號已刪除", "success")
    return redirect("/login")

@app.route("/forgot-password")
def forgot_password_page():
    return render_template("forgot_password.html")


@app.route("/privacy")
def privacy_page():
    return render_template("privacy.html")


@app.route("/terms")
def terms_page():
    return render_template("terms.html")


@app.route("/disclaimer")
def disclaimer_page():
    return render_template("disclaimer.html")

@app.route("/line/liff-bind")
def line_liff_bind_page():
    return render_template(
        "liff_bind.html",
        line_liff_id=LINE_LIFF_ID
    )


@app.route("/line/liff-bind/confirm", methods=["POST"])
def line_liff_bind_confirm_page():
    data = request.get_json() or {}

    line_user_id = data.get("line_user_id", "").strip()
    bind_code = data.get("bind_code", "").strip().upper()

    if not line_user_id:
        return jsonify({
            "success": False,
            "message": "無法取得 LINE 使用者資訊，請重新開啟綁定頁。"
        }), 400

    if not bind_code:
        return jsonify({
            "success": False,
            "message": "找不到綁定碼，請回 Cadouka 個人資料頁重新產生綁定連結。"
        }), 400

    user = get_user_by_line_bind_code(bind_code)

    if not user:
        return jsonify({
            "success": False,
            "message": "找不到這組綁定碼，請回 Cadouka 個人資料頁重新產生。"
        }), 404

    expires_at_text = user["line_bind_code_expires_at"]

    if expires_at_text:
        try:
            expires_at = datetime.strptime(expires_at_text, "%Y-%m-%d %H:%M:%S")

            taiwan_now = datetime.now(timezone(timedelta(hours=8))).replace(tzinfo=None)

            if taiwan_now > expires_at:
                return jsonify({
                    "success": False,
                    "message": "這組綁定碼已過期，請回 Cadouka 個人資料頁重新產生。"
                }), 400
        except:
            return jsonify({
                "success": False,
                "message": "綁定碼狀態異常，請回 Cadouka 個人資料頁重新產生。"
            }), 400

    existing_line_user = get_user_by_line_user_id(line_user_id)

    if existing_line_user and existing_line_user["id"] != user["id"]:
        return jsonify({
            "success": False,
            "message": "這個 LINE 帳號已經綁定其他 Cadouka 帳號，請先解除綁定。"
        }), 409

    bind_line_user_to_account(user["id"], line_user_id)

    safe_add_line_log(
        line_user_id=line_user_id,
        action="bind_line",
        result="success",
        message="LINE 一鍵綁定成功"
    )

    display_name = user["display_name"] or user["username"]

    user_code = user["user_code"] if user.get("user_code") else "-"

    try:
        line_bot_api.push_message(
            line_user_id,
            TextSendMessage(
                text=(
                    "Cadouka 綁定成功！\n\n"
                    f"你的 LINE 已綁定 Cadouka 帳號：{display_name}"
                )
            )
        )
    except Exception as e:
        print("LINE 綁定成功訊息推送失敗：", e)
        traceback.print_exc()

    return jsonify({
        "success": True,
        "message": f"綁定成功！你的 LINE 已綁定 Cadouka 帳號：{display_name}"
    })

# =========================
# Admin Console Routes
# =========================

@app.route("/cdk-console/users/<int:user_id>/membership", methods=["POST"])
@login_required
@admin_required
def admin_update_user_membership_page(user_id):
    membership_level = request.form.get("membership_level", "free").strip()

    if membership_level not in ["free", "pro"]:
        membership_level = "free"

    update_user_membership_level(user_id, membership_level)

    flash("會員等級已更新", "success")
    return redirect(request.referrer or "/cdk-console/users")

@app.route("/cdk-console")
@login_required
@admin_required
def admin_dashboard_page():
    stats = get_admin_overview_stats()

    return render_template(
        "admin_dashboard.html",
        stats=stats
    )


@app.route("/cdk-console/users")
@login_required
@admin_required
def admin_users_page():
    try:
        page = int(request.args.get("page", 1))
    except:
        page = 1

    if page < 1:
        page = 1

    per_page = 20
    total_items = count_admin_users()

    if total_items == 0:
        total_pages = 1
    else:
        total_pages = (total_items + per_page - 1) // per_page

    if page > total_pages:
        page = total_pages

    offset = (page - 1) * per_page

    users = get_admin_users(
        limit=per_page,
        offset=offset
    )

    return render_template(
        "admin_users.html",
        users=users,
        current_page=page,
        total_pages=total_pages,
        total_items=total_items,
        per_page=per_page
    )


@app.route("/cdk-console/cards")
@login_required
@admin_required
def admin_cards_page():
    keyword = request.args.get("keyword", "").strip()

    try:
        page = int(request.args.get("page", 1))
    except:
        page = 1

    if page < 1:
        page = 1

    per_page = 20
    total_items = count_admin_cards(keyword=keyword)

    if total_items == 0:
        total_pages = 1
    else:
        total_pages = (total_items + per_page - 1) // per_page

    if page > total_pages:
        page = total_pages

    offset = (page - 1) * per_page

    cards = get_admin_cards(
        keyword=keyword,
        limit=per_page,
        offset=offset
    )

    card_list = []

    for card in cards:
        card_dict = dict(card)
        card_dict["holding_days"] = calculate_holding_days_for_card(card_dict)
        card_list.append(card_dict)

    return render_template(
        "admin_cards.html",
        cards=card_list,
        keyword=keyword,
        current_page=page,
        total_pages=total_pages,
        total_items=total_items,
        per_page=per_page
    )



def get_log_value(log, key, default=""):
    try:
        if isinstance(log, dict):
            value = log.get(key, default)
        else:
            value = log[key]

        return default if value is None else value
    except:
        return default


def extract_line_log_search_keyword(log):
    raw_keyword = str(get_log_value(log, "raw_keyword", "") or "").strip()
    resolved_keyword = str(get_log_value(log, "resolved_keyword", "") or "").strip()
    action = str(get_log_value(log, "action", "") or "").strip()
    message = str(get_log_value(log, "message", "") or "").strip()

    if raw_keyword:
        return raw_keyword

    if resolved_keyword:
        return resolved_keyword

    if action != "search":
        return ""

    # 舊資料如果沒有 raw_keyword，就從備註回推。
    # 範例：查價成功：116 080 → 116 080，找到 6 筆商品
    if "：" in message:
        keyword_part = message.split("：", 1)[1].split("，", 1)[0].strip()

        if "→" in keyword_part:
            keyword_part = keyword_part.split("→", 1)[0].strip()

        return keyword_part

    return ""


def prepare_line_log_rows(logs):
    log_rows = []

    for log in logs or []:
        try:
            log_dict = dict(log)
        except:
            log_dict = log

        if isinstance(log_dict, dict):
            log_dict["search_keyword"] = extract_line_log_search_keyword(log_dict)

        log_rows.append(log_dict)

    return log_rows

@app.route("/cdk-console/line-logs")
@login_required
@admin_required
def admin_line_logs_page():
    try:
        page = int(request.args.get("page", 1))
    except:
        page = 1

    if page < 1:
        page = 1

    per_page = 20
    total_items = count_admin_line_logs()

    if total_items == 0:
        total_pages = 1
    else:
        total_pages = (total_items + per_page - 1) // per_page

    if page > total_pages:
        page = total_pages

    offset = (page - 1) * per_page

    logs = get_admin_line_logs(
        limit=per_page,
        offset=offset
    )

    logs = prepare_line_log_rows(logs)

    return render_template(
        "admin_line_logs.html",
        logs=logs,
        current_page=page,
        total_pages=total_pages,
        total_items=total_items,
        per_page=per_page
    )


@app.route("/cdk-console/line-logs/export-excel")
@login_required
@admin_required
def admin_line_logs_export_excel_page():
    total_items = count_admin_line_logs()

    logs = get_admin_line_logs(
        limit=total_items if total_items > 0 else 1,
        offset=0
    )

    logs = prepare_line_log_rows(logs)

    wb = Workbook()
    ws = wb.active
    ws.title = "LINE 使用紀錄"

    headers = [
        "時間",
        "用戶編號",
        "用戶",
        "搜尋關鍵字",
        "動作",
        "結果",
        "備註",
        "原始關鍵字",
        "實際搜尋詞",
        "商品數"
    ]

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="DBEAFE")
    header_font = Font(bold=True, color="111827")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for log in logs:
        raw_keyword = get_log_value(log, "raw_keyword", "")
        resolved_keyword = get_log_value(log, "resolved_keyword", "")
        search_keyword = get_log_value(log, "search_keyword", "")

        ws.append([
            get_log_value(log, "created_at", ""),
            get_log_value(log, "user_code", "") or "未綁定",
            get_log_value(log, "display_name", "") or get_log_value(log, "username", "") or "-",
            search_keyword or "-",
            get_log_value(log, "action", "") or "-",
            get_log_value(log, "result", "") or "-",
            get_log_value(log, "message", "") or "-",
            raw_keyword or "-",
            resolved_keyword or "-",
            get_log_value(log, "product_count", "") if get_log_value(log, "product_count", "") != "" else "-"
        ])

    column_widths = {
        "A": 24,
        "B": 14,
        "C": 18,
        "D": 22,
        "E": 20,
        "F": 14,
        "G": 48,
        "H": 22,
        "I": 22,
        "J": 12
    }

    for column_letter, width in column_widths.items():
        ws.column_dimensions[column_letter].width = width

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"cadouka_line_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/cdk-console/line-search-stats")
@login_required
@admin_required
def admin_line_search_stats_page():
    try:
        page = int(request.args.get("page", 1))
    except:
        page = 1

    if page < 1:
        page = 1

    per_page = 20

    popular_keywords = get_line_search_popular_keywords(limit=50)
    no_result_keywords = get_line_search_no_result_keywords(limit=50)

    total_items = count_recent_line_search_logs()

    if total_items == 0:
        total_pages = 1
    else:
        total_pages = (total_items + per_page - 1) // per_page

    if page > total_pages:
        page = total_pages

    offset = (page - 1) * per_page

    recent_logs = get_recent_line_search_logs(
        limit=per_page,
        offset=offset
    )

    return render_template(
        "admin_line_search_stats.html",
        popular_keywords=popular_keywords,
        no_result_keywords=no_result_keywords,
        recent_logs=recent_logs,
        current_page=page,
        total_pages=total_pages,
        total_items=total_items,
        per_page=per_page
    )

@app.route("/cdk-console/search-aliases", methods=["GET", "POST"])
@login_required
@admin_required
def admin_search_aliases_page():
    if request.method == "POST":
        alias_keyword = request.form.get("alias_keyword", "").strip()
        search_keyword = request.form.get("search_keyword", "").strip()
        note = request.form.get("note", "").strip()
        tag_ids = request.form.getlist("tag_ids")

        if len(tag_ids) > 8:
            flash("每筆搜尋別名最多只能選 8 個標籤", "warning")
            return redirect("/cdk-console/search-aliases")

        if not alias_keyword or not search_keyword:
            flash("請輸入俗稱與實際搜尋詞", "warning")
            return redirect("/cdk-console/search-aliases")

        try:
            add_search_alias(
                alias_keyword=alias_keyword,
                search_keyword=search_keyword,
                note=note,
                is_active=1,
                tag_ids=tag_ids
            )
        except Exception as e:
            print("新增搜尋別名錯誤：", e)
            traceback.print_exc()
            flash("新增搜尋別名失敗，可能是俗稱已存在", "warning")

        return redirect("/cdk-console/search-aliases")

    keyword = request.args.get("keyword", "").strip()

    try:
        selected_tag_id = int(request.args.get("tag_id", 0) or 0)
    except:
        selected_tag_id = 0

    try:
        page = int(request.args.get("page", 1))
    except:
        page = 1

    if page < 1:
        page = 1

    per_page = 20
    total_items = count_search_aliases(
        keyword=keyword,
        tag_id=selected_tag_id if selected_tag_id else None
    )

    if total_items == 0:
        total_pages = 1
    else:
        total_pages = (total_items + per_page - 1) // per_page

    if page > total_pages:
        page = total_pages

    offset = (page - 1) * per_page

    aliases = get_all_search_aliases(
        keyword=keyword,
        limit=per_page,
        offset=offset,
        tag_id=selected_tag_id if selected_tag_id else None
    )

    alias_list = []

    for alias in aliases:
        alias_dict = dict(alias)
        created_at_text = format_created_at_taiwan_time(
            alias_dict.get("created_at")
        )
        alias_dict["created_at_text"] = created_at_text

        if created_at_text and len(created_at_text) >= 16:
            alias_dict["created_at_short"] = (
                f"{created_at_text[2:4]}/{created_at_text[5:7]}/{created_at_text[8:10]} "
                f"{created_at_text[11:16]}"
            )
        else:
            alias_dict["created_at_short"] = created_at_text

        alias_list.append(alias_dict)

    tags = get_all_search_tags()
    tag_list = [dict(tag) for tag in tags]

    return render_template(
        "admin_search_aliases.html",
        aliases=alias_list,
        tags=tag_list,
        all_tags=tag_list,
        import_result={
            "total": request.args.get("import_total", ""),
            "created": request.args.get("import_created", ""),
            "updated": request.args.get("import_updated", ""),
            "skipped": request.args.get("import_skipped", ""),
            "errors": request.args.get("import_errors", "")
        },
        keyword=keyword,
        selected_tag_id=selected_tag_id,
        current_page=page,
        total_pages=total_pages,
        total_items=total_items,
        per_page=per_page
    )


# =========================
# Search Alias Import Helpers
# =========================

def normalize_import_header(value):
    return str(value or "").strip().replace("\ufeff", "").lower()


def pick_import_value(row, possible_names):
    normalized_row = {}

    for key, value in row.items():
        normalized_row[normalize_import_header(key)] = "" if value is None else str(value).strip()

    for name in possible_names:
        normalized_name = normalize_import_header(name)

        if normalized_name in normalized_row:
            return normalized_row.get(normalized_name, "").strip()

    return ""


def normalize_import_rows(raw_rows):
    rows = []

    for row in raw_rows or []:
        alias_keyword = pick_import_value(row, [
            "alias_keyword",
            "alias",
            "俗稱",
            "別名",
            "關鍵字",
            "搜尋別名",
            "輸入詞"
        ])

        search_keyword = pick_import_value(row, [
            "search_keyword",
            "search",
            "實際搜尋詞",
            "實際搜尋",
            "搜尋詞",
            "卡號",
            "卡片編號",
            "查詢詞"
        ])

        note = pick_import_value(row, [
            "note",
            "備註",
            "說明"
        ])

        tags = pick_import_value(row, [
            "tags",
            "tag",
            "標籤",
            "分類"
        ])

        if not alias_keyword and not search_keyword:
            continue

        rows.append({
            "alias_keyword": alias_keyword,
            "search_keyword": search_keyword,
            "note": note,
            "tags": tags
        })

    return rows


def decode_uploaded_text_file(file_bytes):
    for encoding in ["utf-8-sig", "utf-8", "cp950", "big5"]:
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue

    return file_bytes.decode("utf-8", errors="ignore")


def parse_search_alias_import_file(uploaded_file):
    filename = (uploaded_file.filename or "").lower()
    file_bytes = uploaded_file.read()

    if filename.endswith(".xlsx"):
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))

        if not rows:
            return []

        headers = [str(cell or "").strip() for cell in rows[0]]
        raw_rows = []

        for values in rows[1:]:
            row = {}

            for index, header in enumerate(headers):
                if not header:
                    continue

                row[header] = values[index] if index < len(values) else ""

            raw_rows.append(row)

        return normalize_import_rows(raw_rows)

    text = decode_uploaded_text_file(file_bytes)
    reader = csv.DictReader(StringIO(text))

    return normalize_import_rows(list(reader))


@app.route("/cdk-console/search-aliases/import", methods=["POST"])
@login_required
@admin_required
def admin_import_search_aliases_page():
    uploaded_file = request.files.get("alias_file")

    if not uploaded_file or not uploaded_file.filename:
        return redirect("/cdk-console/search-aliases")

    try:
        rows = parse_search_alias_import_file(uploaded_file)
        result = bulk_import_search_aliases(rows, max_tags=8)

        return redirect(
            "/cdk-console/search-aliases"
            f"?import_total={result.get('total', 0)}"
            f"&import_created={result.get('created', 0)}"
            f"&import_updated={result.get('updated', 0)}"
            f"&import_skipped={result.get('skipped', 0)}"
            f"&import_errors={result.get('errors', 0)}"
        )

    except Exception as e:
        print("匯入搜尋別名錯誤：", e)
        traceback.print_exc()

        return redirect(
            "/cdk-console/search-aliases"
            "?import_total=0&import_created=0&import_updated=0&import_skipped=0&import_errors=1"
        )


@app.route("/cdk-console/search-tags/add", methods=["POST"])
@login_required
@admin_required
def admin_add_search_tag_page():
    tag_color = request.form.get("tag_color", "#3b82f6").strip()
    tag_name = request.form.get("tag_name", "").strip()

    if not tag_name:
        flash("請輸入標籤名稱", "warning")
        return redirect("/cdk-console/search-aliases")

    try:
        add_search_tag(tag_name=tag_name, tag_color=tag_color)
    except Exception as e:
        print("新增搜尋標籤錯誤：", e)
        traceback.print_exc()
        flash("新增標籤失敗，可能是標籤名稱已存在", "warning")

    return redirect("/cdk-console/search-aliases")


@app.route("/cdk-console/search-tags/<int:tag_id>/update", methods=["POST"])
@login_required
@admin_required
def admin_update_search_tag_page(tag_id):
    tag = get_search_tag_by_id(tag_id)

    if not tag:
        flash("找不到這個標籤", "warning")
        return redirect("/cdk-console/search-aliases")

    tag_color = request.form.get("tag_color", "#3b82f6").strip()
    tag_name = request.form.get("tag_name", "").strip()

    if not tag_name:
        flash("請輸入標籤名稱", "warning")
        return redirect("/cdk-console/search-aliases")

    try:
        update_search_tag(tag_id=tag_id, tag_name=tag_name, tag_color=tag_color)
    except Exception as e:
        print("更新搜尋標籤錯誤：", e)
        traceback.print_exc()
        flash("更新標籤失敗，可能是標籤名稱已存在", "warning")

    return redirect("/cdk-console/search-aliases")


@app.route("/cdk-console/search-tags/<int:tag_id>/delete", methods=["POST"])
@login_required
@admin_required
def admin_delete_search_tag_page(tag_id):
    tag = get_search_tag_by_id(tag_id)

    if not tag:
        flash("找不到這個標籤", "warning")
        return redirect("/cdk-console/search-aliases")

    delete_search_tag(tag_id)
    return redirect("/cdk-console/search-aliases")

@app.route("/cdk-console/search-aliases/<int:alias_id>/update", methods=["POST"])
@login_required
@admin_required
def admin_update_search_alias_page(alias_id):
    alias_keyword = request.form.get("alias_keyword", "").strip()
    search_keyword = request.form.get("search_keyword", "").strip()
    note = request.form.get("note", "").strip()
    tag_ids = request.form.getlist("tag_ids")

    if len(tag_ids) > 8:
        flash("每筆搜尋別名最多只能選 8 個標籤", "warning")
        return redirect("/cdk-console/search-aliases")

    if not alias_keyword or not search_keyword:
        flash("請輸入俗稱與實際搜尋詞", "warning")
        return redirect("/cdk-console/search-aliases")

    alias = get_search_alias_by_id(alias_id)

    if not alias:
        flash("找不到這筆搜尋別名", "warning")
        return redirect("/cdk-console/search-aliases")

    try:
        update_search_alias(
            alias_id=alias_id,
            alias_keyword=alias_keyword,
            search_keyword=search_keyword,
            note=note,
            tag_ids=tag_ids
        )
    except Exception as e:
        print("更新搜尋別名錯誤：", e)
        traceback.print_exc()
        flash("更新搜尋別名失敗，可能是俗稱已存在", "warning")

    return redirect("/cdk-console/search-aliases")

@app.route("/cdk-console/search-aliases/<int:alias_id>/toggle", methods=["POST"])
@login_required
@admin_required
def admin_toggle_search_alias_page(alias_id):
    alias = get_search_alias_by_id(alias_id)

    if not alias:
        flash("找不到這筆搜尋別名", "warning")
        return redirect("/cdk-console/search-aliases")

    try:
        current_active = int(alias["is_active"] or 0)
    except:
        current_active = 0

    new_active = 0 if current_active == 1 else 1

    update_search_alias_active(alias_id, new_active)

    return redirect("/cdk-console/search-aliases")


@app.route("/cdk-console/search-aliases/<int:alias_id>/delete", methods=["POST"])
@login_required
@admin_required
def admin_delete_search_alias_page(alias_id):
    alias = get_search_alias_by_id(alias_id)

    if not alias:
        flash("找不到這筆搜尋別名", "warning")
        return redirect("/cdk-console/search-aliases")

    delete_search_alias(alias_id)

    return redirect("/cdk-console/search-aliases")

# =========================
# Dashboard / Card Routes
# =========================

@app.route("/dashboard")
@login_required
def dashboard_page():
    user_id = current_user_id()

    holding, sold = get_dashboard_full_summary(user_id=user_id)

    holding_cards = holding["total_cards"] or 0
    holding_cost = holding["total_cost"] or 0
    holding_market_value = holding["total_market_value"] or 0
    holding_unrealized_profit = holding["total_unrealized_profit"] or 0

    if holding_cost == 0:
        holding_roi = 0
    else:
        holding_roi = (holding_unrealized_profit / holding_cost) * 100

    sold_cards = sold["total_cards"] or 0
    sold_cost = sold["total_cost"] or 0
    sold_net_revenue = sold["total_net_revenue"] or 0
    sold_realized_profit = sold["total_realized_profit"] or 0

    if sold_cost == 0:
        sold_roi = 0
    else:
        sold_roi = (sold_realized_profit / sold_cost) * 100

    total_cost = holding_cost + sold_cost
    total_profit = holding_unrealized_profit + sold_realized_profit

    if total_cost == 0:
        total_roi = 0
    else:
        total_roi = (total_profit / total_cost) * 100

        # =========================
    # Holding Overview Top 3
    # =========================

    holding_card_list = get_all_cards(
        status="holding",
        keyword=None,
        sort=None,
        user_id=user_id
    )

    holding_card_dicts = []

    for card in holding_card_list:
        card_dict = dict(card)
        holding_card_dicts.append(card_dict)

    top_profit_cards = sorted(
        [
            card for card in holding_card_dicts
            if (card.get("unrealized_profit") or 0) > 0
        ],
        key=lambda card: card.get("unrealized_profit") or 0,
        reverse=True
    )[:3]

    top_loss_cards = sorted(
        [
            card for card in holding_card_dicts
            if (card.get("unrealized_profit") or 0) < 0
        ],
        key=lambda card: card.get("unrealized_profit") or 0
    )[:3]

    return render_template(
        "dashboard.html",

        holding_cards=holding_cards,
        holding_cost=holding_cost,
        holding_market_value=holding_market_value,
        holding_unrealized_profit=holding_unrealized_profit,
        holding_roi=holding_roi,

        sold_cards=sold_cards,
        sold_cost=sold_cost,
        sold_net_revenue=sold_net_revenue,
        sold_realized_profit=sold_realized_profit,
        sold_roi=sold_roi,

        total_cost=total_cost,
        total_profit=total_profit,
        total_roi=total_roi,
        top_profit_cards=top_profit_cards,
        top_loss_cards=top_loss_cards
    )


@app.route("/cards")
@login_required
def card_list_page():
    user_id = current_user_id()

    status = request.args.get("status")
    keyword = request.args.get("keyword", "").strip()
    sort = request.args.get("sort", "").strip()

    try:
        page = int(request.args.get("page", 1))
    except:
        page = 1

    if page < 1:
        page = 1

    per_page = 10

    if status not in ["holding", "sold"]:
        status = None

    allowed_sorts = [
        "date_desc",
        "date_asc",
        "cost_desc",
        "cost_asc",
        "profit_desc",
        "profit_asc",
        "roi_desc",
        "roi_asc"
    ]

    if sort not in allowed_sorts:
        sort = ""

    all_cards = get_all_cards(
        status=status,
        keyword=keyword,
        sort=sort,
        user_id=user_id
    )

    # =========================
    # Summary：用全部符合條件的卡牌計算
    # =========================

    list_total_cards = 0
    list_holding_cards = 0
    list_sold_cards = 0
    list_total_cost = 0
    list_market_or_revenue = 0
    list_total_profit = 0

    for card in all_cards:
        card_dict = dict(card)

        list_total_cards += 1

        buy_price = card_dict.get("buy_price") or 0
        list_total_cost += buy_price

        if card_dict.get("status") == "sold":
            list_sold_cards += 1
            net_revenue = card_dict.get("net_revenue") or 0
            realized_profit = card_dict.get("realized_profit") or 0

            list_market_or_revenue += net_revenue
            list_total_profit += realized_profit
        else:
            list_holding_cards += 1
            current_market_price = card_dict.get("current_market_price") or 0
            unrealized_profit = card_dict.get("unrealized_profit") or 0

            list_market_or_revenue += current_market_price
            list_total_profit += unrealized_profit

    if list_total_cost == 0:
        list_total_roi = 0
    else:
        list_total_roi = (list_total_profit / list_total_cost) * 100

    list_summary = {
        "total_cards": list_total_cards,
        "holding_cards": list_holding_cards,
        "sold_cards": list_sold_cards,
        "total_cost": list_total_cost,
        "market_or_revenue": list_market_or_revenue,
        "total_profit": list_total_profit,
        "total_roi": list_total_roi
    }

    # =========================
    # Pagination：每頁 10 張
    # =========================

    total_items = len(all_cards)

    if total_items == 0:
        total_pages = 1
    else:
        total_pages = (total_items + per_page - 1) // per_page

    if page > total_pages:
        page = total_pages

    start_index = (page - 1) * per_page
    end_index = start_index + per_page

    page_cards = all_cards[start_index:end_index]

    card_list = []

    for card in page_cards:
        card_dict = dict(card)
        card_dict["holding_days"] = calculate_holding_days_for_card(card_dict)
        card_list.append(card_dict)

    latest_price_update_job = get_latest_price_update_job_for_user(user_id)

    return render_template(
        "card_list.html",
        cards=card_list,
        current_status=status,
        keyword=keyword,
        sort=sort,
        list_summary=list_summary,
        latest_price_update_job=latest_price_update_job,
        auto_price_update_mode=request.args.get("auto_price_update", ""),
        auto_price_update_card_id=request.args.get("card_id", ""),

        current_page=page,
        total_pages=total_pages,
        total_items=total_items,
        per_page=per_page
    )

def format_created_at_taiwan_time(created_at):
    if not created_at:
        return ""

    try:
        if isinstance(created_at, datetime):
            dt = created_at
        else:
            created_at_text = str(created_at).split(".")[0]
            dt = datetime.strptime(created_at_text, "%Y-%m-%d %H:%M:%S")

        # 資料庫 CURRENT_TIMESTAMP 通常是 UTC
        dt = dt.replace(tzinfo=timezone.utc)
        taiwan_time = dt.astimezone(timezone(timedelta(hours=8)))

        return taiwan_time.strftime("%Y-%m-%d %H:%M:%S")
    except:
        return str(created_at)

@app.route("/cards/export-excel")
@login_required
def export_cards_excel_page():
    user_id = current_user_id()

    cards = get_all_cards(
        status=None,
        keyword=None,
        sort=None,
        user_id=user_id
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "卡牌倉庫"

    headers = [
        "卡牌名稱",
        "顯示名稱",
        "鑑定卡號",
        "鑑定狀態",
        "狀態",
        "購入日期",
        "購入方式",
        "購入價格",
        "目前市價",
        "未實現損益",
        "未實現 ROI",
        "售出日期",
        "實際收入",
        "已實現損益",
        "已實現 ROI",
        "持有天數",
        "商品網址",
        "備註",
        "建立時間"
    ]

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="DBEAFE")
    header_font = Font(bold=True, color="111827")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for card in cards:
        card_dict = dict(card)

        status_text = "持有中"

        if card_dict.get("status") == "sold":
            status_text = "已售出"
        elif card_dict.get("status") == "holding":
            status_text = "持有中"
        elif card_dict.get("status"):
            status_text = card_dict.get("status")

        holding_days = calculate_holding_days_for_card(card_dict)

        row = [
            card_dict.get("card_name") or "",
            card_dict.get("card_display_name") or "",
            card_dict.get("card_number") or "",
            card_dict.get("grade") or "",
            status_text,
            card_dict.get("buy_date") or "",
            card_dict.get("purchase_method") or "",
            card_dict.get("buy_price") or 0,
            card_dict.get("current_market_price") or 0,
            card_dict.get("unrealized_profit") or 0,
            card_dict.get("roi") or 0,
            card_dict.get("sell_date") or "",
            card_dict.get("net_revenue") or 0,
            card_dict.get("realized_profit") or 0,
            card_dict.get("realized_roi") or 0,
            holding_days,
            card_dict.get("product_url") or "",
            card_dict.get("note") or "",
            format_created_at_taiwan_time(card_dict.get("created_at"))
        ]

        ws.append(row)

    # 欄寬設定
        column_widths = {
        "A": 34,  # 卡牌名稱
        "B": 24,  # 顯示名稱
        "C": 18,  # 鑑定卡號
        "D": 14,  # 鑑定狀態
        "E": 12,  # 狀態
        "F": 14,  # 購入日期
        "G": 18,  # 購入方式
        "H": 14,  # 購入價格
        "I": 14,  # 目前市價
        "J": 16,  # 未實現損益
        "K": 14,  # 未實現 ROI
        "L": 14,  # 售出日期
        "M": 14,  # 實際收入
        "N": 16,  # 已實現損益
        "O": 14,  # 已實現 ROI
        "P": 12,  # 持有天數
        "Q": 42,  # 商品網址
        "R": 30,  # 備註
        "S": 22   # 建立時間
    }

    for column_letter, width in column_widths.items():
        ws.column_dimensions[column_letter].width = width

    # 數字格式
    money_columns = ["H", "I", "J", "M", "N"]

    for column_letter in money_columns:
        for cell in ws[column_letter][1:]:
            cell.number_format = '#,##0'

    percent_columns = ["K", "O"]

    for column_letter in percent_columns:
        for cell in ws[column_letter][1:]:
            cell.number_format = '0.00'

    # 文字垂直置中
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    # 凍結表頭
    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"cadouka_cards_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/cards/add", methods=["GET", "POST"])
@login_required
def add_card_page():
    user_id = current_user_id()

    if request.method == "POST":
        card_name = request.form.get("card_name", "").strip()
        card_display_name = request.form.get("card_display_name", "").strip()
        card_number = request.form.get("card_number", "").strip()
        grade = request.form.get("grade", "").strip()
        purchase_method = request.form.get("purchase_method", "").strip()

        buy_date = request.form.get("buy_date", "").strip()
        buy_price = float(request.form.get("buy_price") or 0)
        current_market_price = float(request.form.get("current_market_price") or 0)

        product_url = request.form.get("product_url", "").strip()
        note = request.form.get("note", "").strip()

        # 目前表單已簡化，所以這些成本固定為 0
        shipping_fee = 0
        tax_fee = 0
        platform_fee = 0
        other_fee = 0

        # LINE/SNKRDUNK 新增時會帶入預設圖片網址。
        # 使用者若上傳圖片，優先使用使用者自己的圖片；沒有上傳才使用預設 image_url。
        image_url = request.form.get("image_url", "").strip()
        uploaded_image_file = request.files.get("card_image")

        if uploaded_image_file and uploaded_image_file.filename:
            try:
                uploaded_image_url = upload_card_image_to_cloudinary(
                    uploaded_image_file,
                    user_id=user_id
                )

                if uploaded_image_url:
                    image_url = uploaded_image_url
            except ValueError as e:
                flash(str(e), "warning")
                return redirect(request.referrer or "/cards/add")

        total_cost = calculate_total_cost(
            buy_price,
            shipping_fee,
            tax_fee,
            platform_fee,
            other_fee
        )

        unrealized_profit, roi = calculate_unrealized_by_buy_price(
            current_market_price,
            buy_price
        )

        card_data = {
            "user_id": user_id,

            "card_name": card_name,
            "card_display_name": card_display_name,
            "card_number": card_number,
            "series_name": "",
            "rarity": "",
            "grade": grade,
            "purchase_method": purchase_method,

            "buy_price": buy_price,
            "shipping_fee": shipping_fee,
            "tax_fee": tax_fee,
            "platform_fee": platform_fee,
            "other_fee": other_fee,

            "total_cost": total_cost,
            "current_market_price": current_market_price,
            "unrealized_profit": unrealized_profit,
            "roi": roi,

            "status": "holding",
            "buy_date": buy_date,
            "image_url": image_url,
            "product_url": product_url,
            "note": note
        }

        add_card(card_data)

        flash("卡牌新增成功", "success")
        return redirect("/cards")

    prefill = {
        "card_name": request.args.get("card_name", "").strip(),
        "card_display_name": request.args.get("card_display_name", "").strip(),
        "card_number": request.args.get("card_number", "").strip(),
        "grade": request.args.get("grade", "").strip(),
        "purchase_method": "",
        "buy_date": request.args.get("buy_date", "").strip(),
        "buy_price": request.args.get("buy_price", "").strip(),
        "current_market_price": request.args.get("current_market_price", "").strip(),
        "image_url": request.args.get("image_url", "").strip(),
        "product_url": request.args.get("product_url", "").strip(),
        "note": request.args.get("note", "").strip()
    }

    return render_template("add_card.html", prefill=prefill)


@app.route("/cards/<int:card_id>")
@login_required
def card_detail_page(card_id):
    user_id = current_user_id()

    card = get_card_by_id(card_id, user_id=user_id)

    if not card:
        return "找不到這張卡牌", 404

    card_dict = dict(card)
    card_dict["holding_days"] = calculate_holding_days_for_card(card_dict)

    # 會員功能
    # user = current_user()
    # is_pro = is_pro_user(user)

    # pro_market_summary = None

    # if is_pro:
    #     pro_market_summary = build_pro_market_summary(
    #         card_dict.get("product_url") or ""
    #     )

    # return render_template(
    #     "card_detail.html",
    #     card=card_dict,
    #     is_pro=is_pro,
    #     membership_level=get_membership_level(user),
    #     pro_market_summary=pro_market_summary
    # )

    user = current_user()

    pro_market_summary = build_pro_market_summary(
        card_dict.get("product_url") or "",
        card_dict
    )

    return render_template(
        "card_detail.html",
        card=card_dict,
        is_pro=True,
        membership_level=get_membership_level(user),
        pro_market_summary=pro_market_summary
    )

@app.route("/pro")
@login_required
def pro_page():
    user = current_user()

    return render_template(
        "pro.html",
        is_pro=is_pro_user(user),
        membership_level=get_membership_level(user)
    )

@app.route("/cards/<int:card_id>/market-history")
@login_required
def card_pro_history_page(card_id):
    user_id = current_user_id()
    user = current_user()

    card = get_card_by_id(card_id, user_id=user_id)

    if not card:
        return "找不到這張卡牌", 404
    
    # 會員功能
    # if not is_pro_user(user):
    #     flash("此功能為 Cadouka Pro 會員功能", "warning")
    #     return redirect(f"/cards/{card_id}")

    selected_grade = request.args.get("grade", "PSA10").strip()

    allowed_grades = get_pro_conditions()

    if selected_grade not in allowed_grades:
        selected_grade = "PSA10"

    card_dict = dict(card)
    card_dict["holding_days"] = calculate_holding_days_for_card(card_dict)

    history_data = build_pro_history_data(
        card_dict,
        selected_grade
    )

    return render_template(
        "card_pro_history.html",
        card=card_dict,
        selected_grade=selected_grade,
        history_data=history_data,
        membership_level=get_membership_level(user),
        is_pro=True
    )


@app.route("/cards/<int:card_id>/edit", methods=["GET", "POST"])
@login_required
def edit_card_page(card_id):
    user_id = current_user_id()

    card = get_card_by_id(card_id, user_id=user_id)

    if not card:
        return "找不到這張卡牌", 404

    if request.method == "POST":
        card_name = request.form.get("card_name", "").strip()
        card_display_name = request.form.get("card_display_name", "").strip()
        card_number = request.form.get("card_number", "").strip()
        grade = request.form.get("grade", "").strip()
        purchase_method = request.form.get("purchase_method", "").strip()

        buy_date = request.form.get("buy_date", "").strip()
        buy_price = float(request.form.get("buy_price") or 0)
        current_market_price = float(request.form.get("current_market_price") or 0)

        product_url = request.form.get("product_url", "").strip()
        note = request.form.get("note", "").strip()

        # 目前表單已簡化，所以這些成本固定為 0
        shipping_fee = 0
        tax_fee = 0
        platform_fee = 0
        other_fee = 0

        # 編輯時預設保留原本圖片；使用者若上傳新圖片，優先替換成使用者自己的圖片。
        image_url = card["image_url"] or ""
        uploaded_image_file = request.files.get("card_image")

        if uploaded_image_file and uploaded_image_file.filename:
            try:
                uploaded_image_url = upload_card_image_to_cloudinary(
                    uploaded_image_file,
                    user_id=user_id
                )

                if uploaded_image_url:
                    image_url = uploaded_image_url
            except ValueError as e:
                flash(str(e), "warning")
                return redirect(request.referrer or f"/cards/{card_id}/edit")

        total_cost = calculate_total_cost(
            buy_price,
            shipping_fee,
            tax_fee,
            platform_fee,
            other_fee
        )

        unrealized_profit, roi = calculate_unrealized_by_buy_price(
            current_market_price,
            buy_price
        )

        card_data = {
            "card_name": card_name,
            "card_display_name": card_display_name,
            "card_number": card_number,
            "series_name": "",
            "rarity": "",
            "grade": grade,
            "purchase_method": purchase_method,

            "buy_price": buy_price,
            "shipping_fee": shipping_fee,
            "tax_fee": tax_fee,
            "platform_fee": platform_fee,
            "other_fee": other_fee,

            "total_cost": total_cost,
            "current_market_price": current_market_price,
            "unrealized_profit": unrealized_profit,
            "roi": roi,

            "buy_date": buy_date,
            "image_url": image_url,
            "product_url": product_url,
            "note": note
        }

        update_card(card_id, card_data, user_id=user_id)

        # 如果這張卡已經售出，購入價格改變後，
        # 已實現損益與已實現 ROI 也要重新計算。
        if card["status"] == "sold":
            net_revenue = card["net_revenue"] or card["sell_price"] or 0

            realized_profit, realized_roi = calculate_realized_by_buy_price(
                net_revenue,
                buy_price
            )

            sell_data = {
                "sell_price": card["sell_price"] or net_revenue,
                "sell_fee": card["sell_fee"] or 0,
                "sell_shipping_fee": card["sell_shipping_fee"] or 0,
                "sell_other_fee": card["sell_other_fee"] or 0,
                "net_revenue": net_revenue,
                "realized_profit": realized_profit,
                "realized_roi": realized_roi,
                "sell_date": card["sell_date"] or ""
            }

            mark_card_as_sold(card_id, sell_data, user_id=user_id)

        flash("卡牌資料已更新", "success")
        return redirect("/cards")

    return render_template("edit_card.html", card=card)


@app.route("/cards/<int:card_id>/delete", methods=["POST"])
@login_required
def delete_card_page(card_id):
    user_id = current_user_id()

    card = get_card_by_id(card_id, user_id=user_id)

    if not card:
        return "找不到這張卡牌", 404

    delete_card(card_id, user_id=user_id)

    flash("卡牌已刪除", "success")
    return redirect("/cards")

@app.route("/cards/bulk-delete", methods=["POST"])
@login_required
def bulk_delete_cards_page():
    user_id = current_user_id()

    card_ids = request.form.getlist("card_ids")

    if not card_ids:
        flash("請先選擇要刪除的卡牌", "warning")
        return redirect(request.referrer or "/cards")

    deleted_count = 0

    for card_id in card_ids:
        try:
            delete_card(int(card_id), user_id=user_id)
            deleted_count += 1
        except Exception as e:
            print("批量刪除錯誤：", e)
            traceback.print_exc()
            continue

    flash(f"已刪除 {deleted_count} 張卡牌", "success")
    return redirect("/cards")

@app.route("/cards/<int:card_id>/sell", methods=["GET", "POST"])
@login_required
def sell_card_page(card_id):
    user_id = current_user_id()

    card = get_card_by_id(card_id, user_id=user_id)

    if not card:
        return "找不到這張卡牌", 404

    if request.method == "POST":
        sell_date = request.form.get("sell_date", "").strip()

        # 現在這個欄位代表「實際收入」
        sell_price = float(request.form.get("sell_price") or 0)

        # 目前售出表單已簡化，所以售出成本固定為 0
        sell_fee = 0
        sell_shipping_fee = 0
        sell_other_fee = 0

        # 實際收入 = 使用者填寫的金額
        net_revenue = sell_price

        buy_price = card["buy_price"] or 0

        realized_profit, realized_roi = calculate_realized_by_buy_price(
            net_revenue,
            buy_price
        )

        sell_data = {
            "sell_price": sell_price,
            "sell_fee": sell_fee,
            "sell_shipping_fee": sell_shipping_fee,
            "sell_other_fee": sell_other_fee,
            "net_revenue": net_revenue,
            "realized_profit": realized_profit,
            "realized_roi": realized_roi,
            "sell_date": sell_date
        }

        mark_card_as_sold(card_id, sell_data, user_id=user_id)

        if card["status"] == "sold":
            flash("售出資料已更新", "success")
        else:
            flash("已標記為已售出", "success")

        return redirect(f"/cards/{card_id}")

    return render_template("sell_card.html", card=card)


@app.route("/cards/<int:card_id>/unsell", methods=["POST"])
@login_required
def unsell_card_page(card_id):
    user_id = current_user_id()

    card = get_card_by_id(card_id, user_id=user_id)

    if not card:
        return "找不到這張卡牌", 404

    mark_card_as_holding(card_id, user_id=user_id)

    flash("已標記回持有中", "success")
    return redirect(request.referrer or f"/cards/{card_id}")

@app.route("/cards/<int:card_id>/refresh-price", methods=["POST"])
@login_required
def refresh_single_card_price_page(card_id):
    """
    Render Free / Web-only 版本：
    單張更新不建立 worker 任務，而是導回卡牌倉庫後由前端 AJAX 分批處理。
    """
    return redirect(f"/cards?auto_price_update=single&card_id={int(card_id)}")


@app.route("/cards/refresh-all-prices", methods=["POST"])
@login_required
def refresh_all_card_prices_page():
    """
    Render Free / Web-only 版本：
    全部更新由前端分批呼叫 API，不需要 background worker。
    """
    return redirect("/cards?auto_price_update=all")


@app.route("/cards/refresh-prices/start", methods=["POST"])
@login_required
def start_web_batch_price_update_page():
    user_id = current_user_id()
    data = request.get_json(silent=True) or {}

    mode = str(data.get("mode") or "all").strip()

    if mode not in ["all", "single"]:
        mode = "all"

    try:
        card_id = int(data.get("card_id") or 0)
    except:
        card_id = 0

    # 清掉太久沒有完成的任務，避免使用者關掉頁面後永遠被鎖住。
    try:
        fail_stale_price_update_jobs(
            user_id=user_id,
            stale_minutes=PRICE_UPDATE_STALE_MINUTES
        )
    except Exception as e:
        print("清理過期市價任務失敗：", e)

    if mode == "all":
        if has_active_price_update_job(user_id, job_type="all"):
            return jsonify({
                "success": False,
                "message": "你已經有一個市價更新正在進行中，請稍後再試。"
            }), 409

        if has_recent_price_update_job(
            user_id,
            job_type="all",
            within_minutes=PRICE_UPDATE_LOCK_MINUTES
        ):
            return jsonify({
                "success": False,
                "message": "10 分鐘內已更新過全部市價，請稍後再試。"
            }), 429

    if mode == "single":
        if card_id <= 0:
            return jsonify({
                "success": False,
                "message": "找不到要更新的卡牌。"
            }), 400

        if has_active_price_update_job(user_id, card_id=card_id, job_type="single"):
            return jsonify({
                "success": False,
                "message": "這張卡已經在更新中，請稍後再試。"
            }), 409

    card_ids = collect_price_update_card_ids(
        user_id=user_id,
        mode=mode,
        card_id=card_id if mode == "single" else None
    )

    if not card_ids:
        if mode == "all":
            message = "目前沒有需要更新的卡牌。可能是 6 小時內已更新過，或卡牌尚未設定商品網址。"
        else:
            message = "這張卡目前無法更新，可能是已售出、找不到卡牌，或缺少商品網址。"

        return jsonify({
            "success": False,
            "message": message
        }), 200

    job_id = create_price_update_job(
        user_id=user_id,
        job_type=mode,
        card_id=card_id if mode == "single" else None,
        message="Web 分批更新準備中"
    )

    now_text = get_taiwan_now_text()
    mark_price_update_job_running(job_id, now_text)
    update_price_update_job_progress(
        job_id,
        total_count=len(card_ids),
        updated_count=0,
        skipped_count=0,
        failed_count=0,
        message="前端分批更新中"
    )

    return jsonify({
        "success": True,
        "job_id": job_id,
        "mode": mode,
        "card_ids": card_ids,
        "total_count": len(card_ids),
        "batch_size": PRICE_UPDATE_BATCH_SIZE,
        "message": "開始分批更新市價"
    })


@app.route("/cards/refresh-prices/batch", methods=["POST"])
@login_required
def run_web_batch_price_update_page():
    user_id = current_user_id()
    data = request.get_json(silent=True) or {}

    try:
        job_id = int(data.get("job_id") or 0)
    except:
        job_id = 0

    raw_card_ids = data.get("card_ids") or []
    is_last = bool(data.get("is_last"))

    if job_id <= 0:
        return jsonify({
            "success": False,
            "message": "找不到更新任務。"
        }), 400

    job = get_price_update_job(job_id)

    if not job or int(get_card_value(job, "user_id", 0) or 0) != int(user_id):
        return jsonify({
            "success": False,
            "message": "找不到更新任務。"
        }), 404

    if get_card_value(job, "status", "") != "running":
        return jsonify({
            "success": False,
            "message": "這個更新任務不是執行中狀態。"
        }), 409

    card_ids = []

    for raw_card_id in raw_card_ids:
        try:
            card_id_int = int(raw_card_id)
        except:
            continue

        if card_id_int > 0 and card_id_int not in card_ids:
            card_ids.append(card_id_int)

    card_ids = card_ids[:max(1, PRICE_UPDATE_BATCH_SIZE)]

    batch_results = []
    batch_updated = 0
    batch_skipped = 0
    batch_failed = 0

    for card_id in card_ids:
        result = process_one_card_market_price_update(card_id, user_id)
        batch_results.append(result)

        if result.get("status") == "updated":
            batch_updated += 1
        elif result.get("status") == "skipped":
            batch_skipped += 1
        else:
            batch_failed += 1

    current_updated = int(get_card_value(job, "updated_count", 0) or 0) + batch_updated
    current_skipped = int(get_card_value(job, "skipped_count", 0) or 0) + batch_skipped
    current_failed = int(get_card_value(job, "failed_count", 0) or 0) + batch_failed
    total_count = int(get_card_value(job, "total_count", 0) or 0)
    done_count = current_updated + current_skipped + current_failed

    message = f"已處理 {done_count}/{total_count} 張"

    update_price_update_job_progress(
        job_id,
        updated_count=current_updated,
        skipped_count=current_skipped,
        failed_count=current_failed,
        message=message
    )

    final_status = "running"

    if is_last or done_count >= total_count:
        if current_updated == 0 and current_failed > 0 and current_skipped == 0:
            final_status = "failed"
            final_message = "市價更新失敗，請稍後再試。"
        else:
            final_status = "done"
            final_message = f"更新完成：成功 {current_updated} 張、略過 {current_skipped} 張、失敗 {current_failed} 張。"

        finish_price_update_job(
            job_id,
            status=final_status,
            message=final_message,
            finished_at=get_taiwan_now_text(),
            updated_count=current_updated,
            skipped_count=current_skipped,
            failed_count=current_failed,
            total_count=total_count
        )

        message = final_message

    return jsonify({
        "success": True,
        "job_id": job_id,
        "status": final_status,
        "total_count": total_count,
        "updated_count": current_updated,
        "skipped_count": current_skipped,
        "failed_count": current_failed,
        "done_count": done_count,
        "message": message,
        "batch_results": batch_results
    })


# =========================
# LINE Callback / Image Tools
# =========================

@app.route("/callback", methods=["POST"])
def callback():
    signature = request.headers.get("X-Line-Signature")
    body = request.get_data(as_text=True)

    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        print("簽章驗證失敗，請檢查 Channel secret")
        abort(400)
    except Exception as e:
        print("callback 錯誤：", e)
        traceback.print_exc()
        abort(500)

    return "OK"


@app.route("/crop-image")
def crop_image():
    image_url = request.args.get("url", "")

    if not image_url:
        return "Missing image url", 400

    try:
        image_url = unquote(image_url)

        request_obj = req.Request(image_url, headers=headers)

        with req.urlopen(request_obj) as response:
            image_bytes = response.read()

        cropped_image = crop_white_border(image_bytes)

        return send_file(
            cropped_image,
            mimetype="image/jpeg"
        )

    except Exception as e:
        print("圖片裁切錯誤：", e)
        traceback.print_exc()
        return "Image processing error", 500

@handler.add(MessageEvent, message=TextMessage)
def handle_message(event):
    card_id = event.message.text.strip()
    line_user_id = event.source.user_id

    # =========================
    # LINE Account Binding Commands
    # =========================

    if card_id.startswith("綁定 "):
        bind_code = card_id.replace("綁定 ", "", 1).strip().upper()

        if not bind_code:
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text="請輸入綁定碼，例如：綁定 A8K29Q")
            )
            return

        user = get_user_by_line_bind_code(bind_code)

        if not user:
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text="找不到這組綁定碼，請確認是否輸入正確。")
            )
            return

        expires_at_text = user["line_bind_code_expires_at"]

        if expires_at_text:
            try:
                expires_at = datetime.strptime(expires_at_text, "%Y-%m-%d %H:%M:%S")

                taiwan_now = datetime.now(timezone(timedelta(hours=8))).replace(tzinfo=None)

                if taiwan_now > expires_at:
                    line_bot_api.reply_message(
                        event.reply_token,
                        TextSendMessage(text="這組綁定碼已過期，請回 Cadouka 個人資料頁重新產生。")
                    )
                    return
            except:
                line_bot_api.reply_message(
                    event.reply_token,
                    TextSendMessage(text="綁定碼狀態異常，請回 Cadouka 個人資料頁重新產生。")
                )
                return

        existing_line_user = get_user_by_line_user_id(line_user_id)

        if existing_line_user and existing_line_user["id"] != user["id"]:
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text="這個 LINE 帳號已經綁定其他 Cadouka 帳號，請先解除綁定。")
            )
            return

        bind_line_user_to_account(user["id"], line_user_id)

        safe_add_line_log(
            line_user_id=line_user_id,
            action="bind_line",
            result="success",
            message="LINE 備用綁定碼綁定成功"
        )

        display_name = user["display_name"] or user["username"]

        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text=f"綁定成功！你的 LINE 已綁定 Cadouka 帳號：{display_name}")
        )
        return

    if card_id == "綁定狀態":
        user = get_user_by_line_user_id(line_user_id)

        if not user:
            safe_add_line_log(
                line_user_id=line_user_id,
                action="binding_status",
                result="unbound",
                message="查看綁定狀態：尚未綁定"
            )

            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text="你目前尚未綁定 Cadouka 帳號。")
            )
            return

        safe_add_line_log(
            line_user_id=line_user_id,
            action="binding_status",
            result="success",
            message="查看綁定狀態"
        )

        display_name = user["display_name"] or user["username"]

        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text=f"你目前已綁定 Cadouka 帳號：{display_name}")
        )
        return

    # 第一次輸入解除綁定，只做確認，不直接解除
    if card_id == "解除綁定":
        user = get_user_by_line_user_id(line_user_id)

        if not user:
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text="你目前沒有綁定任何 Cadouka 帳號。")
            )
            return

        line_bot_api.reply_message(
            event.reply_token,
            TemplateSendMessage(
                alt_text="解除綁定確認",
                template=ConfirmTemplate(
                    text=(
                        "確定要解除綁定嗎？\n"
                        "解除後，將無法從 LINE 直接加入 Cadouka。"
                    ),
                    actions=[
                        MessageAction(
                            label="取消",
                            text="取消解除綁定"
                        ),
                        MessageAction(
                            label="確認解除",
                            text="確認解除綁定"
                        )
                    ]
                )
            )
        )
        return

    if card_id == "取消解除綁定":
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text="已取消解除綁定。")
        )
        return

    # 第二次確認，才真正解除綁定
    if card_id == "確認解除綁定":
        user = get_user_by_line_user_id(line_user_id)

        if not user:
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text="你目前沒有綁定任何 Cadouka 帳號。")
            )
            return

        safe_add_line_log(
            line_user_id=line_user_id,
            action="unbind_line",
            result="success",
            message="解除 LINE 綁定"
        )

        unbind_line_user(line_user_id)

        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text="已解除 LINE 與 Cadouka 帳號的綁定。")
        )
        return

    if card_id == "功能":
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(
                text="Cadouka 常用功能：",
                quick_reply=create_main_quick_reply()
            )
        )
        return

    if card_id == "使用教學":
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(
                text="Cadouka 使用教學：",
                quick_reply=create_tutorial_quick_reply()
            )
        )
        return

    if card_id == "教學：如何綁定":
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(
                text=(
                    "【建立 Cadouka 帳號】\n"
                    "1. 進入 Cadouka 網站。\n"
                    "2. 註冊帳號並登入。\n\n"

                    "【LINE 一鍵綁定】\n"
                    "1. 登入 Cadouka 網站。\n"
                    "2. 點右上角選單或帳號名稱，進入「個人資料」。\n"
                    "3. 在 LINE 帳號綁定區按「產生綁定連結」。\n"
                    "4. 按「LINE 一鍵綁定」。\n"
                    "5. 系統會開啟 LINE 並自動完成綁定。\n\n"

                    "如果一鍵綁定無法使用，也可以使用備用綁定碼：\n"
                    "複製畫面上的「綁定 XXXXXX」，到 LINE 貼上送出即可完成綁定。"
                ),
                quick_reply=create_tutorial_quick_reply()
            )
        )
        return

    if card_id == "教學：如何查價":
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(
                text=(
                    "【LINE 查價】\n"
                    "在 LINE 直接輸入關鍵字，例如：\n"
                    "Pikachu\n"
                    "系列名稱：M2 SV2a...\n"
                    "MUR\n"
                    "卡牌左下角編號：XXX XXX"
                ),
                quick_reply=create_tutorial_quick_reply()
            )
        )
        return
    
    if card_id == "教學：加入倉庫":
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(
                text=(
                    "【加入卡牌倉庫】\n"
                    "在 LINE 查價結果中，按「加入 Cadouka」，即可把該商品加入你的卡牌倉庫。\n"
                    "加入後可到網站編輯：\n"
                    "卡牌名稱、鑑定卡號、購入價格...等\n\n"

                    "【市價更新】\n"
                    "如果卡牌有商品網址，可以更新市價。\n"
                    "持有中的卡牌才會更新市價，已售出的卡牌不會更新。\n"
                    "卡牌倉庫中，持有中卡牌右側的更新符號可以更新單張卡牌市價。\n"
                    "上方的「更新市價」可以批量更新持有中的卡牌。"
                ),
                quick_reply=create_tutorial_quick_reply()
            )
        )
        return

    if card_id == "教學：常用指令":
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(
                text=(
                    "【常用指令】\n"
                    "綁定狀態：查看目前 LINE 是否已綁定 Cadouka 帳號\n"
                    "解除綁定：解除目前 LINE 綁定\n"
                    "使用教學：查看 Cadouka 使用方式\n"
                    "功能：查看功能"
                ),
                quick_reply=create_tutorial_quick_reply()
            )
        )
        return

    # =========================
    # Search SNKRDUNK Products
    # =========================

    try:
        resolved_keyword = resolve_search_alias(card_id)
        products = search_products(resolved_keyword)

        if not products:
            safe_add_line_log(
                line_user_id=line_user_id,
                action="search",
                result="no_result",
                message=f"查無商品：{card_id} → {resolved_keyword}",
                raw_keyword=card_id,
                resolved_keyword=resolved_keyword,
                product_count=0
            )

            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text="查無商品，請換一個卡號試試看。")
            )
            return

        user_products[line_user_id] = products

        safe_add_line_log(
            line_user_id=line_user_id,
            action="search",
            result="success",
            message=f"查價成功：{card_id} → {resolved_keyword}，找到 {len(products)} 筆商品",
            raw_keyword=card_id,
            resolved_keyword=resolved_keyword,
            product_count=len(products)
        )

        flex_message = create_product_image_grid_messages(products)

        line_bot_api.reply_message(
            event.reply_token,
            flex_message
        )

    except Exception as e:
        print("搜尋錯誤：", e)
        traceback.print_exc()

        safe_add_line_log(
            line_user_id=line_user_id,
            action="search",
            result="failed",
            message="搜尋時發生錯誤",
            raw_keyword=card_id,
            resolved_keyword=card_id,
            product_count=0
        )

        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text="搜尋時發生錯誤，請稍後再試。")
        )


@handler.add(PostbackEvent)
def handle_postback(event):
    line_user_id = event.source.user_id
    data = event.postback.data


    print("收到 Postback：", data, flush=True)

    try:
        params = parse_qs(data)

        action = params.get("action", [""])[0]

        print("Postback action：", action, flush=True)

        products = user_products.get(line_user_id)


        print("products =", products, flush=True)
        print("products count =", len(products) if products else 0, flush=True)

        if not products:
            safe_add_line_log(
                line_user_id=line_user_id,
                action=action or "postback",
                result="failed",
                message="找不到搜尋結果，請重新輸入"
            )

            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text="找不到搜尋結果，請重新輸入。")
            )
            return

        if action == "select":

                print("進入 select", flush=True)

                index = int(params.get("index", [0])[0])
                selected_grade = params.get("grade", ["PSA10"])[0] or "PSA10"

                print("index =", index, "selected_grade =", selected_grade, flush=True)


                if index >= len(products):
                    line_bot_api.reply_message(
                        event.reply_token,
                        TextSendMessage(text="商品選擇錯誤，請重新搜尋。")
                    )
                    return

                product = products[index]
                product_url = product["url"]
                product_id = get_product_id(product_url)

                line_user = get_user_by_line_user_id(line_user_id)
                line_is_pro = bool(line_user and is_pro_user(line_user))

                if selected_grade in ["A", "B"] and not line_is_pro:
                    line_bot_api.reply_message(
                        event.reply_token,
                        TextSendMessage(text="A / B 行情圖卡為 Cadouka Pro 會員功能。")
                    )
                    return

                price_url = build_sales_history_url(
                    product_id,
                    condition=selected_grade,
                    page=1,
                    per_page=20
                )

                prices = getprice(price_url)

                print("成交資料筆數：", len(prices or []), flush=True)

                jpy_rate = get_jpy_spot_sell()

                filename = generate_market_card_image(
                    product=product,
                    prices=prices,
                    selected_grade=selected_grade,
                    jpy_rate=jpy_rate
                )


                print("產生圖卡檔名：", filename, flush=True)

                base_url = get_base_url()
                card_image_url = f"{base_url}/static/generated/{filename}"

                market_flex_message = create_market_image_card_flex(
                    product=product,
                    card_image_url=card_image_url,
                    product_index=index,
                    selected_grade=selected_grade,
                    is_pro=line_is_pro
                )

                safe_add_line_log(
                    line_user_id=line_user_id,
                    action="select_product",
                    result="success",
                    message=f"點選商品查看 {selected_grade} 圖卡"
                )


                print("準備回傳 LINE Flex", flush=True)

                line_bot_api.reply_message(
                    event.reply_token,
                    market_flex_message
                )

                print("LINE Flex 已送出", flush=True)

                return

        if action == "history":
            index = int(params.get("index", [0])[0])
            selected_grade = params.get("grade", ["PSA10"])[0]

            if index >= len(products):
                line_bot_api.reply_message(
                    event.reply_token,
                    TextSendMessage(text="商品選擇錯誤，請重新搜尋。")
                )
                return

            product = products[index]
            product_url = product["url"]

            line_user = get_user_by_line_user_id(line_user_id)
            line_is_pro = bool(line_user and is_pro_user(line_user))

            if selected_grade in ["A", "B"] and not line_is_pro:
                line_bot_api.reply_message(
                    event.reply_token,
                    TextSendMessage(text="A / B 歷史成交為 Cadouka Pro 會員功能。")
                )
                return

            product_id = get_product_id(product_url)
            price_url = build_sales_history_url(
                product_id,
                condition=selected_grade,
                page=1,
                per_page=20
            )

            prices = getprice(price_url)
            jpy_rate = get_jpy_spot_sell()

            if line_is_pro:
                history_display_limit = 10
            else:
                history_display_limit = 5

            history_flex_message = create_history_flex(
                product,
                prices,
                selected_grade,
                jpy_rate,
                index,
                display_limit=history_display_limit
            )

            safe_add_line_log(
                line_user_id=line_user_id,
                action="history",
                result="success",
                message=f"查看 {selected_grade} 歷史成交"
            )

            line_bot_api.reply_message(
                event.reply_token,
                history_flex_message
            )
            return

        if action == "add_card":
            index = int(params.get("index", [0])[0])
            selected_grade = params.get("grade", [""])[0]

            if index >= len(products):
                line_bot_api.reply_message(
                    event.reply_token,
                    TextSendMessage(text="商品選擇錯誤，請重新搜尋。")
                )
                return

            user = get_user_by_line_user_id(line_user_id)

            if selected_grade in ["A", "B"] and (not user or not is_pro_user(user)):
                safe_add_line_log(
                    line_user_id=line_user_id,
                    action="add_card",
                    result="failed",
                    message="非 Pro，無法加入 A / B 卡牌"
                )

                line_bot_api.reply_message(
                    event.reply_token,
                    TextSendMessage(text="A / B 加入卡牌為 Cadouka Pro 會員功能。")
                )
                return

            if not user:
                safe_add_line_log(
                    line_user_id=line_user_id,
                    action="add_card",
                    result="failed",
                    message="未綁定，無法加入卡牌"
                )

                line_bot_api.reply_message(
                    event.reply_token,
                    TextSendMessage(
                        text=(
                            "你尚未綁定 Cadouka 帳號。\n\n"
                            "請先登入 Cadouka，進入個人資料頁產生 LINE 綁定連結，"
                            "再完成 LINE 一鍵綁定。"
                        ),
                        quick_reply=create_unbound_quick_reply()
                    )
                )
                return

            product = products[index]
            product_name = product["name"] if product.get("name") else "未命名商品"
            image_url = product.get("image") or ""

            product_url = product["url"]

            product_id = get_product_id(product_url)

            if selected_grade:
                price_url = build_sales_history_url(product_id, condition=selected_grade)
            else:
                price_url = build_sales_history_url(product_id)

            prices = getprice(price_url)
            jpy_rate = get_jpy_spot_sell()

            current_market_price = calculate_market_price_from_prices(
                prices,
                jpy_rate
            )

            buy_price = 0
            total_cost = 0

            unrealized_profit, roi = calculate_unrealized_by_buy_price(
                current_market_price,
                buy_price
            )

            card_data = {
                "user_id": user["id"],

                "card_name": product_name,
                "card_display_name": "",
                "card_number": "",
                "series_name": "",
                "rarity": "",
                "grade": selected_grade,
                "purchase_method": "",

                "buy_price": buy_price,
                "shipping_fee": 0,
                "tax_fee": 0,
                "platform_fee": 0,
                "other_fee": 0,

                "total_cost": total_cost,
                "current_market_price": current_market_price,
                "unrealized_profit": unrealized_profit,
                "roi": roi,

                "status": "holding",
                "buy_date": "",
                "image_url": image_url,
                "product_url": product_url,
                "note": ""
            }

            add_card(card_data)

            safe_add_line_log(
                line_user_id=line_user_id,
                action="add_card",
                result="success",
                message=f"加入 {selected_grade or '未填'} 卡牌"
            )

            display_name = user["display_name"] or user["username"]

            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(
                    text=(
                        "已新增到卡牌倉庫！\n"
                        f"帳號：{display_name}\n"
                        f"卡牌：{product_name}\n"
                        f"鑑定狀態：{selected_grade or '未填'}\n"
                        f"目前市價：NT${current_market_price:,}\n\n"
                        "之後可到網站編輯購入價格、購入日期、鑑定卡號與鑑定狀態。"
                    )
                )
            )
            return

    except Exception as e:
        print("選擇商品錯誤：", e)
        traceback.print_exc()

        safe_add_line_log(
            line_user_id=line_user_id,
            action="postback",
            result="failed",
            message="查詢時發生錯誤"
        )

        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text="查詢時發生錯誤，請稍後再試。")
        )


if __name__ == "__main__":
    app.run(port=5000, debug=True)